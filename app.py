"""
Atlantic Packaging — Invoice Tools  v3.0
• Invoice Splitter   : Split batch PDFs into individual invoice PDFs
• Invoice Matcher    : Match invoices with POs, flatten & merge into single PDFs
• Invoice Codifier   : GL / Cost-Centre stamp on invoice PDFs
"""

import streamlit as st
import pdfplumber
from pypdf import PdfReader, PdfWriter
from reportlab.pdfgen import canvas
from io import BytesIO
import zipfile
import re
import copy
import json
import pandas as pd
from datetime import date, datetime, timedelta
import openpyxl
from PIL import Image, ImageDraw

# ── Optional OCR support ──────────────────────────────────────────────────────
try:
    import pytesseract
    from pdf2image import convert_from_bytes
    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False

# ── Optional pikepdf support (needed for Invoice Matcher) ─────────────────────
try:
    import pikepdf
    PIKEPDF_AVAILABLE = True
except ImportError:
    PIKEPDF_AVAILABLE = False

# ─────────────────────────────────────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Atlantic — Invoice Tools",
    layout="wide",
    page_icon="📄",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────────────────────────────────────
# DEFAULT DATA
# ─────────────────────────────────────────────────────────────────────────────
DEFAULT_PROVEEDORES = [
    {"prefijo": "ML", "vendor": "0101000430", "cc": "ML01"},
    {"prefijo": "EV", "vendor": "0101002430", "cc": "EV01"},
    {"prefijo": "MV", "vendor": "0101002220", "cc": "MV01"},
    {"prefijo": "MF", "vendor": "0101000430", "cc": "MF01"},
    {"prefijo": "MS", "vendor": "0101000430", "cc": "MS01"},
    {"prefijo": "MD", "vendor": "0101002430", "cc": "EV01"},
]

DEFAULT_GL_CODES = [
    {"codigo": "CM1023", "gl": "300052"}, {"codigo": "CM1026", "gl": "300051"},
    {"codigo": "CM1030", "gl": "300052"}, {"codigo": "CM1036", "gl": "300052"},
    {"codigo": "KGLB35", "gl": "300051"}, {"codigo": "LB2032", "gl": "300051"},
    {"codigo": "LB2035", "gl": "300051"}, {"codigo": "LB2042", "gl": "300051"},
    {"codigo": "LB2052", "gl": "300051"}, {"codigo": "PL2030", "gl": "300051"},
    {"codigo": "PRCM23", "gl": "300052"}, {"codigo": "WT56",   "gl": "300054"},
    {"codigo": "PRCM30", "gl": "300052"}, {"codigo": "CAWT33", "gl": "300054"},
    {"codigo": "KGLB42", "gl": "300051"}, {"codigo": "MDLU36", "gl": "300052"},
    {"codigo": "CAWT41", "gl": "300054"}, {"codigo": "CSLB42", "gl": "300051"},
    {"codigo": "003514", "gl": "300041"}, {"codigo": "003022", "gl": "300041"},
    {"codigo": "003502", "gl": "300041"}, {"codigo": "003024", "gl": "300041"},
    {"codigo": "DTCM23", "gl": "300052"}, {"codigo": "NDCM30", "gl": "300052"},
    {"codigo": "DTCM30", "gl": "300052"}, {"codigo": "LB2056", "gl": "300051"},
    {"codigo": "CM1033", "gl": "300052"}, {"codigo": "CAWT36", "gl": "300054"},
    {"codigo": "MDLU23", "gl": "300052"}, {"codigo": "CAWT25", "gl": "300054"},
    {"codigo": "003660", "gl": "300041"}, {"codigo": "003500", "gl": "300041"},
    {"codigo": "003771", "gl": "300041"}, {"codigo": "003021", "gl": "300041"},
    {"codigo": "001119", "gl": "300041"}, {"codigo": "002480", "gl": "300041"},
    {"codigo": "003675", "gl": "300041"}, {"codigo": "003501", "gl": "300041"},
    {"codigo": "003727", "gl": "300041"}, {"codigo": "003728", "gl": "300041"},
    {"codigo": "003729", "gl": "300041"}, {"codigo": "001777", "gl": "300041"},
    {"codigo": "002020", "gl": "300041"}, {"codigo": "002728", "gl": "300041"},
    {"codigo": "001912", "gl": "300041"}, {"codigo": "003366", "gl": "300041"},
    {"codigo": "003166", "gl": "300041"}, {"codigo": "002481", "gl": "300041"},
    {"codigo": "003691", "gl": "300041"}, {"codigo": "002488", "gl": "300041"},
    {"codigo": "003607", "gl": "300041"}, {"codigo": "002901", "gl": "300041"},
    {"codigo": "NDCM23", "gl": "300052"}, {"codigo": "WT36",   "gl": "300054"},
    {"codigo": "WT26",   "gl": "300054"}, {"codigo": "WT42",   "gl": "300054"},
    {"codigo": "WT31",   "gl": "300054"},
]

DEFAULT_USERS = ["ROC", "MLE", "PD"]
VENDOR_EXCEPCION = "0101000390"

# Control Facturas — Otros Proveedores: default workflow states. Any value
# already used in imported/legacy data is added on top of this list rather
# than replacing it, so nothing typed before gets orphaned.
CF_DEFAULT_ESTADOS = [
    "Pendiente de codificación",
    "Codificada - pendiente de sello",
    "Sellada - enviada a aprobación",
    "Aprobada",
    "Pagada",
    "Con problema",
]

# ─────────────────────────────────────────────────────────────────────────────
# SESSION STATE INIT
# ─────────────────────────────────────────────────────────────────────────────
def init_state():
    defaults = {
        "active_module":       None,
        "proveedores":         copy.deepcopy(DEFAULT_PROVEEDORES),
        "gl_codes":            copy.deepcopy(DEFAULT_GL_CODES),
        "usuarios":            DEFAULT_USERS.copy(),
        "processed":           [],
        "stamp_x":             281,
        "stamp_y_top":         594,
        "stamp_w":             230,
        "stamp_h":             82,
        # Matcher state
        "matcher_results":     None,
        "matcher_zip":         None,
        "matcher_upload_key":  0,
        # Splitter state
        "splitter_results":    [],   # [{"filename", "pdf_bytes", "invoice_no", "cc", "bol", "pages", "warning", "ocr_used"}]
        "splitter_zip":        None,
        "splitter_upload_key": 0,
        "splitter_rotation":   0,    # rotation applied to scanned (OCR) batches
        # AP Audit state
        "audit_results":       None,
        "audit_data_count":    0,
        # Payment Packager state
        "payment_result":          None,
        "payment_batches":         [],   # [{"label", "files": [{"name", "bytes"}]}]
        "payment_batch_form_key":  0,
        # Reconciliation state
        "recon_results":           None,
        "recon_zip":               None,
        "recon_upload_key":        0,
        "recon_statement_total":   None,
        "recon_grand_total":       None,
        "recon_check_po":          True,
        # Control Facturas — Otros Proveedores (non-Atlantic) state
        "cf_facturas":              [],   # invoice reception/coding registry
        "cf_proveedores":           [],   # vendor catalogue
        "cf_reglas":                [],   # per-vendor GL/CC special rules
        "cf_responsables":          [],   # editable list of people who can be assigned
        "cf_estados":               CF_DEFAULT_ESTADOS.copy(),  # editable status list
        "cf_next_id":               1,
        "cf_stamp_x":               40,
        "cf_stamp_y":               40,
        "cf_stamp_w":               230,
        "cf_stamp_h":               110,
        "cf_stamped_result":        None,
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

init_state()

# ─────────────────────────────────────────────────────────────────────────────
# ── SPLITTER FUNCTIONS ────────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────

def _splitter_get_cc(raw_prefix: str) -> str:
    """
    Map the raw 2-char prefix from Customer Order No. to the CC used in filename.
    Looks up the vendor table: prefijo -> cc field -> take first 2 chars.
    e.g.  ML -> ML01 -> "ML"
          MD -> EV01 -> "EV"
    Falls back to the raw prefix if not found.
    """
    if not raw_prefix:
        return "??"
    key = raw_prefix.upper().strip()
    for row in st.session_state.proveedores:
        if str(row.get("prefijo", "")).upper().strip() == key:
            cc_val = str(row.get("cc", ""))
            return cc_val[:2].upper() if len(cc_val) >= 2 else cc_val.upper()
    return key  # fallback: use raw prefix as-is


def _splitter_extract_invoice_no(lines: list) -> str | None:
    """
    Invoice number appears on the line JUST BEFORE 'INVOICE No/No DE FACTURE'.
    Also handles it trailing on the same line.
    """
    for i, line in enumerate(lines):
        if "INVOICE NO" in line.upper() and "FACTURE" in line.upper():
            # Number at end of same line (allow trailing OCR noise like a
            # stray "-" or "=" picked up from a table border on scanned batches)
            m = re.search(r"(\d{7,10})[^\d]*$", line)
            if m:
                return m.group(1)
            # Number on any of the 4 preceding lines
            for j in range(max(0, i - 4), i):
                m = re.fullmatch(r"\d{6,10}", lines[j].strip())
                if m:
                    return lines[j].strip()
            break
    return None


def _splitter_extract_order_no(lines: list) -> str | None:
    """
    Customer Order No. appears in a line like:
      635108  100  ML11465  Dec 18, 2025  778713917 ...
    Returns the full order number e.g. 'ML11465'.
    """
    for line in lines:
        # \|?\s* tolerates a stray "|" the way OCR sometimes misreads the
        # table's vertical grid line on scanned batches
        m = re.search(r"\d{6}\s+\d{2,3}\s+\|?\s*([A-Za-z]{2}\d{4,7})\b", line, re.IGNORECASE)
        if m:
            return m.group(1).upper()
    return None


def _splitter_extract_bol(lines: list) -> str | None:
    """
    Bill of Lading rules:
    • Two-line BOL — the second line contains the value to use:
        Layout A:  F0002775  1  CAWT25  ...  (line 1)
                   354192  195.331  MSF  4824  LB  ML11465  (line 2 ← USE THIS)
        Layout B:  F0002776  3  PRCM30  ...  (line 1)
                   24286651  606.06  MSF  18120  LB  ML11673  (line 2 ← USE THIS)
      Both layouts: line 2 starts with digits (5–9), then decimal/integer, then MSF.
      The only difference was digit length (6 vs 8) — now covered by \d{5,9}.
    • Single-line BOL:
        N0089112  8  LB2035  35# EnviroLiner ...  (alphanumeric ← USE THIS)
    Always use the LAST (second) BOL value when two exist.
    For multi-item invoices, use the BOL from the FIRST item only.
    """
    # Priority: second-line BOL — identifier followed by decimal+MSF
    # The identifier can be:
    #   • Pure digits   : 354192, 24286651  (layouts A & B)
    #   • Letter+digits : C249328           (layout C — letter-prefixed BOL)
    # [A-Z]? makes the leading letter optional, covering all three layouts.
    for line in lines:
        m = re.match(r"^([A-Za-z]{0,2}\d{5,9})\s+[\d,]+\.[\d]+\s+MSF\b", line.strip(), re.IGNORECASE)
        if m:
            return m.group(1).upper()

    # Fallback: alphanumeric single-line BOL (e.g. C0012857, N0089112)
    # Product code after qty can be alphabetic (CAWT25) OR numeric (001912) —
    # only require: letter + 5-9 digits, space, qty digits, space
    for line in lines:
        m = re.match(r"^([A-Za-z]\d{5,9})\s+\d+\s", line.strip(), re.IGNORECASE)
        if m:
            return m.group(1).upper()

    return None


def _splitter_render_page_image(pdf_bytes: bytes, page_no: int, dpi: int = 200):
    """Render one 1-based page of a PDF to a PIL image, or None if OCR deps
    are unavailable or rendering fails."""
    if not OCR_AVAILABLE:
        return None
    try:
        images = convert_from_bytes(pdf_bytes, first_page=page_no, last_page=page_no, dpi=dpi)
    except Exception:
        return None
    return images[0] if images else None


def _splitter_suggest_rotation(pdf_bytes: bytes, sample_pages: int = 3) -> int:
    """
    Best-effort guess at the page rotation needed for a scanned batch —
    scanners commonly output pages sideways. Scores each candidate angle
    (90/270 checked before 0/180, since a sideways scan is the far more
    common case) by OCR-ing a few sample pages and rewarding a page whose
    text actually matches the invoice-number pattern; this is only a
    *default* — the UI shows thumbnails of all four candidates so the
    user confirms or overrides it visually before anything is split.
    """
    reader = PdfReader(BytesIO(pdf_bytes))
    n = min(sample_pages, len(reader.pages))
    if n == 0:
        return 0

    scores = {90: 0.0, 270: 0.0, 0: 0.0, 180: 0.0}
    for page_no in range(1, n + 1):
        img = _splitter_render_page_image(pdf_bytes, page_no, dpi=150)
        if img is None:
            continue
        for angle in scores:
            rimg = img.rotate(angle, expand=True) if angle else img
            try:
                text = pytesseract.image_to_string(rimg)
            except Exception:
                text = ""
            lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
            if _splitter_extract_invoice_no(lines):
                scores[angle] += 20
            t = text.upper()
            if "INVOICE" in t:
                scores[angle] += 3
            if "FACTURE" in t:
                scores[angle] += 3
            if "ATLANTIC" in t:
                scores[angle] += 1
    return max(scores, key=scores.get)


def split_batch_pdf(pdf_bytes: bytes, rotation: int = 0, progress_callback=None) -> list:
    """
    Split a multi-invoice Atlantic batch PDF into individual invoices.

    Handles both regular (text-layer) batch PDFs and scanned batches with
    no embedded text — those fall back to OCR automatically. `rotation`
    (0/90/180/270) corrects pages that came out of the scanner sideways:
    it's applied before OCR *and* baked into the output PDFs so the split
    invoices open upright. progress_callback(page_no, total_pages), when
    given, is called for each page while OCR is running (it's slow).

    Returns list of dicts:
      { filename, pdf_bytes, invoice_no, cc, bol, pages (1-based),
        source_pages, page_count, warning, ocr_used }
    """
    reader = PdfReader(BytesIO(pdf_bytes))

    with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
        page_texts = [p.extract_text() or "" for p in pdf.pages]

    total_pages = len(page_texts)
    needs_ocr = OCR_AVAILABLE and len("".join(page_texts[:3]).strip()) < 30

    if needs_ocr:
        page_texts = []
        for page_idx in range(total_pages):
            if progress_callback:
                progress_callback(page_idx + 1, total_pages)
            img = _splitter_render_page_image(pdf_bytes, page_idx + 1, dpi=200)
            if img is not None and rotation:
                img = img.rotate(rotation, expand=True)
            page_texts.append(pytesseract.image_to_string(img) if img is not None else "")

    invoices = []   # accumulated invoice dicts
    current  = None # {"number", "cc_raw", "bol", "pages": [0-based idx]}

    for page_idx, text in enumerate(page_texts):
        lines = [ln.strip() for ln in text.splitlines() if ln.strip()]

        is_continuation = (
            "...Continued from previous page" in text
            or "Continued from previous page" in text
        )

        if is_continuation and current is not None:
            current["pages"].append(page_idx)
            continue

        inv_no = _splitter_extract_invoice_no(lines)

        if inv_no:
            if current is not None:
                invoices.append(current)
            order_no = _splitter_extract_order_no(lines)
            cc_raw   = order_no[:2].upper() if order_no else None
            bol      = _splitter_extract_bol(lines)
            current  = {
                "number":  inv_no,
                "cc_raw":  cc_raw,
                "bol":     bol,
                "pages":   [page_idx],
            }
        elif current is not None:
            # Page without a clear invoice marker — attach to current
            current["pages"].append(page_idx)

    if current is not None:
        invoices.append(current)

    # Build output list
    results = []
    for inv in invoices:
        writer = PdfWriter()
        for p in inv["pages"]:
            page = reader.pages[p]
            if needs_ocr and rotation:
                # `rotation` is expressed in PIL's counter-clockwise
                # convention (it's what corrects the OCR preview image);
                # pypdf's Page.rotate() is clockwise, so it needs the
                # opposite angle to land on the same upright orientation.
                page.rotate((-rotation) % 360)
            writer.add_page(page)
        buf = BytesIO()
        writer.write(buf)

        inv_no = inv["number"]
        cc     = _splitter_get_cc(inv["cc_raw"]) if inv["cc_raw"] else "??"
        bol    = inv["bol"] or "??"
        warn   = None
        if inv["cc_raw"] is None:
            warn = "⚠️ Customer Order No. not detected — CC set to '??'"
        if inv["bol"] is None:
            warn = (warn or "") + "  ⚠️ Bill of Lading not detected — BOL set to '??'"
        if needs_ocr:
            warn = (warn or "") + "  📷 Read via OCR — please double-check the fields"

        results.append({
            "filename":     f"{inv_no} {cc} {bol}.pdf",
            "pdf_bytes":    buf.getvalue(),
            "invoice_no":   inv_no,
            "cc":           cc,
            "bol":          bol,
            "source_pages": [p + 1 for p in inv["pages"]],
            "page_count":   len(inv["pages"]),
            "warning":      warn,
            "ocr_used":     needs_ocr,
        })

    return results


def make_splitter_zip(results: list) -> bytes:
    buf = BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for item in results:
            zf.writestr(item["filename"], item["pdf_bytes"])
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────────────────────────────────────
# ── CODIFIER FUNCTIONS ────────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────

def get_vendor_cc(prefix: str):
    prefix = prefix.upper()
    for row in st.session_state.proveedores:
        if str(row["prefijo"]).upper() == prefix:
            return str(row["vendor"]), str(row["cc"])
    return None, None


def get_gl(product_code: str) -> str | None:
    if not product_code:
        return None
    code = product_code.upper().strip()
    for row in st.session_state.gl_codes:
        if str(row["codigo"]).upper().strip() == code:
            return str(row["gl"])
    return None


def extract_invoice_data(pdf_bytes: bytes, filename: str = "") -> dict:
    result = {
        "invoice_no": None, "customer_order": None, "cc_prefix": None,
        "product_code": None, "is_six": False, "raw_lines": [],
        "ocr_used": False, "error": None,
    }
    try:
        with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
            if not pdf.pages:
                result["error"] = "Empty PDF"
                return result
            text = pdf.pages[0].extract_text() or ""

        if len(text.strip()) < 50:
            if OCR_AVAILABLE:
                try:
                    images = convert_from_bytes(pdf_bytes, first_page=1, last_page=1, dpi=200)
                    text = pytesseract.image_to_string(images[0])
                    result["ocr_used"] = True
                except Exception as ocr_err:
                    result["error"] = f"OCR failed: {ocr_err}"
                    return result
            else:
                result["error"] = "Scanned PDF — OCR not available"
                return result

        lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
        result["raw_lines"] = lines
        text_upper = text.upper()

        for i, line in enumerate(lines):
            if "INVOICE NO" in line.upper() and "FACTURE" in line.upper():
                m = re.search(r"(\d{7,10})\s*$", line)
                if m:
                    result["invoice_no"] = m.group(1)
                    result["is_six"] = result["invoice_no"].startswith("6")
                    break
                for j in range(max(0, i - 5), i):
                    if re.fullmatch(r"\d{6,10}", lines[j]):
                        result["invoice_no"] = lines[j]
                        result["is_six"] = result["invoice_no"].startswith("6")
                        break
                break

        if not result["invoice_no"] and filename:
            m = re.match(r"(\d{6,10})", filename)
            if m:
                result["invoice_no"] = m.group(1)
                result["is_six"] = result["invoice_no"].startswith("6")

        for line in lines:
            m = re.search(r"\d{6}\s+\d{2,3}\s+\|?\s*([A-Za-z]{2}\d{4,7})\b", line, re.IGNORECASE)
            if m:
                result["customer_order"] = m.group(1).upper()
                result["cc_prefix"] = m.group(1)[:2].upper()
                break

        known_codes = sorted(
            [str(r["codigo"]).upper() for r in st.session_state.gl_codes],
            key=len, reverse=True,
        )
        for code in known_codes:
            if re.search(r"\b" + re.escape(code) + r"\b", text_upper):
                result["product_code"] = code
                break

    except Exception as e:
        result["error"] = str(e)

    return result


def create_stamp(user, vendor, cc, gl, coding_date, page_w, page_h, rotation=0):
    sw = st.session_state.stamp_w
    sh = st.session_state.stamp_h
    margin = 18
    date_str = (coding_date.strftime("%d/%m/%Y")
                if hasattr(coding_date, "strftime") else str(coding_date))
    stamp_lines = [
        f"POSTED BY: {user}",
        f"VENDOR: {vendor}",
        f"CC: {cc}  |  GL: {gl}",
        f"DATE: {date_str}",
    ]
    packet = BytesIO()
    c = canvas.Canvas(packet, pagesize=(page_w, page_h))
    if rotation in (90, 270):
        disp_w, disp_h = page_h, page_w
        disp_cx = disp_w / 2
        disp_cy = disp_h - margin - sh / 2
        if rotation == 90:
            cx_pdf, cy_pdf = page_w - disp_cy, disp_cx
            rot_angle = 90
        else:
            cx_pdf, cy_pdf = disp_cy, page_w - disp_cx
            rot_angle = 90
        c.saveState()
        c.translate(cx_pdf, cy_pdf)
        c.rotate(rot_angle)
        lx, ly = -sw / 2, -sh / 2
        c.setStrokeColorRGB(0.85, 0.0, 0.0)
        c.setFillColorRGB(1.0, 1.0, 1.0)
        c.setLineWidth(1.8)
        c.rect(lx, ly, sw, sh, fill=1)
        c.setFillColorRGB(0.85, 0.0, 0.0)
        c.setFont("Helvetica-Bold", 8.5)
        line_h = sh / 5.2
        tx, ty = lx + 10, ly + sh - line_h
        for i, line in enumerate(stamp_lines):
            c.drawString(tx, ty - i * line_h, line)
        c.restoreState()
    else:
        sx = st.session_state.stamp_x
        sy_top = st.session_state.stamp_y_top
        sy_bot = sy_top - sh
        c.setStrokeColorRGB(0.85, 0.0, 0.0)
        c.setFillColorRGB(1.0, 1.0, 1.0)
        c.setLineWidth(1.8)
        c.rect(sx, sy_bot, sw, sh, fill=1)
        c.setFillColorRGB(0.85, 0.0, 0.0)
        c.setFont("Helvetica-Bold", 8.5)
        line_h = sh / 5.2
        tx, ty = sx + 10, sy_bot + sh - line_h
        for i, line in enumerate(stamp_lines):
            c.drawString(tx, ty - i * line_h, line)
    c.save()
    packet.seek(0)
    return packet.read()


def stamp_pdf(original_bytes, stamp_bytes):
    reader = PdfReader(BytesIO(original_bytes))
    stamp_reader = PdfReader(BytesIO(stamp_bytes))
    stamp_page = stamp_reader.pages[0]
    writer = PdfWriter()
    first = reader.pages[0]
    first.merge_page(stamp_page)
    writer.add_page(first)
    for page in reader.pages[1:]:
        writer.add_page(page)
    out = BytesIO()
    writer.write(out)
    out.seek(0)
    return out.read()


def process_one(original_bytes, user, vendor, cc, gl, coding_date):
    reader = PdfReader(BytesIO(original_bytes))
    page = reader.pages[0]
    pw = float(page.mediabox.width)
    ph = float(page.mediabox.height)
    rotation = int(page.get("/Rotate", 0) or 0)
    stamp_bytes = create_stamp(user, vendor, cc, gl, coding_date, pw, ph, rotation)
    return stamp_pdf(original_bytes, stamp_bytes)


def make_zip(items):
    buf = BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for item in items:
            zf.writestr(item["filename"], item["pdf_bytes"])
    buf.seek(0)
    return buf.read()


def _get_processed_zip() -> bytes:
    items = st.session_state.processed
    if not items:
        return b""
    sig = tuple(item["filename"] for item in items)
    if st.session_state.get("_proc_zip_sig") == sig and st.session_state.get("_proc_zip"):
        return st.session_state["_proc_zip"]
    z = make_zip(items)
    st.session_state["_proc_zip"]     = z
    st.session_state["_proc_zip_sig"] = sig
    return z


# ─────────────────────────────────────────────────────────────────────────────
# ── INVOICE MATCHER FUNCTIONS ─────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────

def extract_po_from_invoice_name(filename: str) -> str | None:
    """
    Invoice filename format: '{invoice_no} {cost_center} {PO_number}.pdf'
    e.g. '82196530 ML V0020978.pdf'  →  'V0020978'
    """
    name = re.sub(r'\.pdf$', '', filename, flags=re.IGNORECASE).strip()
    parts = name.split(' ')
    if len(parts) >= 3:
        return parts[2].strip()
    return None


def flatten_pdf(pdf_bytes: bytes) -> bytes:
    if PIKEPDF_AVAILABLE:
        try:
            inp = BytesIO(pdf_bytes)
            out = BytesIO()
            with pikepdf.open(inp, suppress_warnings=True) as pdf:
                pdf.save(out)
            out.seek(0)
            return out.read()
        except Exception:
            pass

    if OCR_AVAILABLE:
        try:
            from pdf2image import convert_from_bytes
            pages = convert_from_bytes(pdf_bytes, dpi=120)
            out = BytesIO()
            if len(pages) == 1:
                pages[0].save(out, format="PDF")
            else:
                pages[0].save(out, format="PDF", save_all=True,
                              append_images=pages[1:])
            out.seek(0)
            return out.read()
        except Exception:
            pass

    return pdf_bytes


def merge_two_pdfs(bytes1: bytes, bytes2: bytes) -> bytes:
    if PIKEPDF_AVAILABLE:
        out = BytesIO()
        with pikepdf.Pdf.new() as merged:
            with pikepdf.open(BytesIO(bytes1), suppress_warnings=True) as p1:
                merged.pages.extend(p1.pages)
            with pikepdf.open(BytesIO(bytes2), suppress_warnings=True) as p2:
                merged.pages.extend(p2.pages)
            merged.save(out)
        out.seek(0)
        return out.read()
    else:
        writer = PdfWriter()
        for b in (bytes1, bytes2):
            reader = PdfReader(BytesIO(b))
            for page in reader.pages:
                writer.add_page(page)
        out = BytesIO()
        writer.write(out)
        out.seek(0)
        return out.read()


def run_matching(invoice_files: list, po_files: list,
                 progress_callback=None) -> dict:
    po_lookup = {}
    for f in po_files:
        key = re.sub(r'\.pdf$', '', f.name, flags=re.IGNORECASE).strip().upper()
        po_lookup[key] = f.read()

    used_po_keys = set()
    matched = []
    pending = []
    total = len(invoice_files)

    for i, inv_file in enumerate(invoice_files):
        inv_bytes = inv_file.read()
        fname = inv_file.name

        if progress_callback:
            progress_callback(i, total, fname)

        po_id = extract_po_from_invoice_name(fname)

        if po_id is None:
            pending.append({
                "invoice_name": fname,
                "po_id":        "—",
                "reason":       "Invalid filename format (needs at least 3 space-separated parts)",
                "inv_bytes":    inv_bytes,
            })
            continue

        po_key = po_id.upper()
        if po_key in po_lookup:
            po_bytes = po_lookup[po_key]
            used_po_keys.add(po_key)
            flat_inv = flatten_pdf(inv_bytes)
            flat_po  = flatten_pdf(po_bytes)
            merged   = merge_two_pdfs(flat_inv, flat_po)
            matched.append({
                "invoice_name": fname,
                "po_name":      f"{po_id}.pdf",
                "po_id":        po_id,
                "merged_bytes": merged,
            })
        else:
            pending.append({
                "invoice_name": fname,
                "po_id":        po_id,
                "reason":       f"No PO file found for '{po_id}'",
                "inv_bytes":    inv_bytes,
            })

    unmatched_po = [
        name for name in po_lookup
        if name not in used_po_keys
    ]

    return {
        "matched":       matched,
        "pending":       pending,
        "unmatched_po":  unmatched_po,
    }


def make_matcher_zip(results: dict) -> bytes:
    buf = BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for item in results["matched"]:
            zf.writestr(f"matched/{item['invoice_name']}", item["merged_bytes"])
        for item in results["pending"]:
            zf.writestr(f"pending/{item['invoice_name']}", item["inv_bytes"])
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────────────────────────────────────
# ── COURU CODE FUNCTIONS ──────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────

def extract_couru_data(pdf_bytes: bytes, filename: str = "") -> dict:
    result = {
        "invoice_no": None, "date": None,
        "gl": None, "cc": None, "subtotal": None, "error": None,
    }
    try:
        with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
            if not pdf.pages:
                result["error"] = "Empty PDF"
                return result
            full_text = ""
            for page in pdf.pages:
                full_text += (page.extract_text() or "") + "\n"
            text_upper = full_text.upper()
            lines = [ln.strip() for ln in full_text.splitlines() if ln.strip()]

            # Invoice number
            for i, line in enumerate(lines):
                if "INVOICE NO" in line.upper() and "FACTURE" in line.upper():
                    m = re.search(r"(\d{7,10})\s*$", line)
                    if m:
                        result["invoice_no"] = m.group(1)
                        break
                    for j in range(max(0, i - 5), i):
                        if re.fullmatch(r"\d{6,10}", lines[j]):
                            result["invoice_no"] = lines[j]
                            break
                    break
            if not result["invoice_no"] and filename:
                m = re.match(r"(\d{6,10})", filename)
                if m:
                    result["invoice_no"] = m.group(1)

            # Invoice date — last date on the customer-order header line
            date_pat = r'(\d{1,2}\s+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\w*\s+\d{4})'
            for line in lines:
                if re.search(r'\d{6}\s+\d{2,3}\s+[A-Za-z]{2}\d+', line, re.IGNORECASE):
                    dates = re.findall(date_pat, line, re.IGNORECASE)
                    if dates:
                        result["date"] = dates[-1]
                        break
            if not result["date"]:
                m = re.search(date_pat, full_text, re.IGNORECASE)
                if m:
                    result["date"] = m.group(1)

            # CC from customer order prefix
            for line in lines:
                m = re.search(r"\d{6}\s+\d{2,3}\s+\|?\s*([A-Za-z]{2}\d{4,7})\b", line, re.IGNORECASE)
                if m:
                    _, cc = get_vendor_cc(m.group(1)[:2].upper())
                    result["cc"] = cc
                    break

            # GL from product codes
            known_codes = sorted(
                [str(r["codigo"]).upper() for r in st.session_state.gl_codes],
                key=len, reverse=True,
            )
            if known_codes:
                _pat = r"\b(?:" + "|".join(re.escape(c) for c in known_codes) + r")\b"
                _m = re.search(_pat, text_upper)
                if _m:
                    result["gl"] = get_gl(_m.group(0))

            # Subtotal (total before taxes) — find amount near "Sub Total" label
            for i, line in enumerate(lines):
                if re.search(r'\bsub\s*total\b', line, re.IGNORECASE):
                    amounts = re.findall(r'[\d,]+\.\d{2}', line)
                    if amounts:
                        result["subtotal"] = amounts[0].replace(",", "")
                        break
                    for j in range(max(0, i - 4), i):
                        amounts = re.findall(r'^([\d,]+\.\d{2})$', lines[j])
                        if amounts:
                            result["subtotal"] = amounts[0].replace(",", "")
                            break
                    if result["subtotal"]:
                        break
                    for j in range(i + 1, min(len(lines), i + 3)):
                        amounts = re.findall(r'([\d,]+\.\d{2})', lines[j])
                        if amounts:
                            result["subtotal"] = amounts[0].replace(",", "")
                            break
                    break

    except Exception as e:
        result["error"] = str(e)
    return result


# ─────────────────────────────────────────────────────────────────────────────
# ── AP AUDIT VALIDATION FUNCTIONS ─────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────

def _parse_date_str(s: str) -> "date | None":
    """Parse a date string in multiple formats → date object."""
    if not s:
        return None
    s = s.strip()
    for fmt in (
        "%d %b %Y", "%d %B %Y",
        "%d-%b-%Y", "%d-%B-%Y",
        "%d-%b-%y", "%d-%B-%y",
        "%d/%m/%Y", "%m/%d/%Y",
        "%d/%m/%y", "%m/%d/%y",
        "%Y-%m-%d", "%Y/%m/%d",
        "%d.%m.%Y",
    ):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def calc_due_date(invoice_date) -> "date | None":
    """
    Quincena rule: invoice_date + 60 days → round UP to the 15th or 30th.
    The second quincena is always the 30th (never 31).
    """
    if isinstance(invoice_date, str):
        invoice_date = _parse_date_str(invoice_date)
    if not invoice_date:
        return None
    base = invoice_date + timedelta(days=60)
    if base.day <= 15:
        return date(base.year, base.month, 15)
    return date(base.year, base.month, 30)


def extract_invoice_date(pdf_bytes: bytes) -> "date | None":
    """
    Extract the INVOICE DATE from an Atlantic invoice PDF.
    Tries: 1) 'INVOICE DATE' label search  2) last date on the customer-order line.
    """
    date_pat = r'(\d{1,2}\s+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\w*\s+\d{4})'
    try:
        with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
            full_text = ""
            for page in pdf.pages:
                full_text += (page.extract_text() or "") + "\n"
            lines = [ln.strip() for ln in full_text.splitlines() if ln.strip()]

            # Method 1: find 'INVOICE DATE' label and grab the adjacent date
            for i, line in enumerate(lines):
                if re.search(r'INVOICE\s+DATE', line, re.IGNORECASE):
                    window = " ".join(lines[i:min(i + 3, len(lines))])
                    m = re.search(date_pat, window, re.IGNORECASE)
                    if m:
                        return _parse_date_str(m.group(1))

            # Method 2: last date on the customer-order header line
            for line in lines:
                if re.search(r'\d{6}\s+\d{2,3}\s+[A-Za-z]{2}\d+', line, re.IGNORECASE):
                    dates = re.findall(date_pat, line, re.IGNORECASE)
                    if dates:
                        return _parse_date_str(dates[-1])
    except Exception:
        pass
    return None


def parse_audit_report(pdf_bytes: bytes) -> dict:
    """
    Parse the A/P Voucher Audit Listing PDF (Crystal Reports format).

    All pages are merged into a single coordinate space (y-offset per page)
    so invoice blocks that span a page boundary are handled correctly.
    Each block runs from [APINV_y, next_APINV_y) — no cross-invoice contamination.
    Invoice date = date on the same line as the APINV/invoice-number row;
    due date = the next date found in the rows below it (yk+6..yk+45).
    Total is located by the 'Total:' label; falls back to max amount in block.
    """
    records = {}
    known_ccs = {r["cc"].upper() for r in st.session_state.proveedores}

    _ISO_PAT    = re.compile(r'^\d{4}-\d{2}-\d{2}$')
    _OTHER_PATS = [
        re.compile(r'^\d{1,2}-(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)-\d{2,4}$', re.I),
        re.compile(r'^\d{2}/\d{2}/\d{4}$'),
        re.compile(r'^\d{4}/\d{2}/\d{2}$'),
        re.compile(r'^\d{2}\.\d{2}\.\d{4}$'),
    ]
    # Non-anchored versions for searching inside a reconstructed character run
    _DATE_SEARCH_PATS = [
        re.compile(r'\d{4}-\d{2}-\d{2}'),
        re.compile(r'\d{1,2}-(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)-\d{2,4}', re.I),
        re.compile(r'\d{2}/\d{2}/\d{4}'),
        re.compile(r'\d{4}/\d{2}/\d{2}'),
        re.compile(r'\d{2}\.\d{2}\.\d{4}'),
    ]

    def _is_date_word(text: str) -> "date | None":
        t = text.strip()
        if _ISO_PAT.match(t):
            return _parse_date_str(t)
        for p in _OTHER_PATS:
            if p.match(t):
                return _parse_date_str(t)
        return None

    def _row_date(ys: list) -> "date | None":
        """First date word found scanning the given rows top-to-bottom."""
        for y in ys:
            for w in sorted(row_map.get(y, []), key=lambda w: w["x0"]):
                d = _is_date_word(w["text"])
                if d:
                    return d
        return None

    def _chars_date(top_lo: float, top_hi: float, x_lo, x_hi) -> "date | None":
        """
        Read a date directly from the character stream within a y/x box.

        Crystal Reports can overflow a long Party name into the date
        column's x-range; both text runs then share the same x0 per
        character cell, and pdfplumber's word merge garbles them into one
        string (e.g. "0M2L6-08-12" instead of "2026-08-12"). Rebuilding
        the cell from characters and picking, at each x0, the glyph most
        likely to be a genuine date character (digit > separator > other)
        recovers the real date even when the overflow text itself
        contains a hyphen at that same position.
        """
        if x_lo is None or x_hi is None:
            return None
        cell = [c for c in all_chars
                if top_lo <= c["top"] <= top_hi and x_lo - 5 <= c["x0"] < x_hi]
        if not cell:
            return None

        def _priority(t: str) -> int:
            if t.isdigit():
                return 2
            if t in "-/.":
                return 1
            return 0

        by_x = {}
        for c in cell:
            x0 = round(c["x0"], 1)
            t = c["text"]
            pr = _priority(t)
            prev = by_x.get(x0)
            if prev is None or pr > prev[1]:
                by_x[x0] = (t, pr)
        s = "".join(v[0] for _, v in sorted(by_x.items()))
        for p in _DATE_SEARCH_PATS:
            m = p.search(s)
            if m:
                return _parse_date_str(m.group(0))
        return None

    try:
        # ── Merge all pages into one coordinate space ────────────────────────────
        all_words: list = []
        all_chars: list = []
        y_offset = 0.0
        with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
            for page in pdf.pages:
                pw = page.extract_words(
                    x_tolerance=3, y_tolerance=3,
                    keep_blank_chars=False, use_text_flow=False,
                )
                for w in pw:
                    nw = dict(w)
                    nw["top"]    = w["top"]    + y_offset
                    nw["bottom"] = w["bottom"] + y_offset
                    all_words.append(nw)
                for ch in page.chars:
                    nc = dict(ch)
                    nc["top"] = ch["top"] + y_offset
                    all_chars.append(nc)
                y_offset += page.height

        if not all_words:
            return records

        # Date column x-range, from the "Invoice Date / Due Date" header
        # (both sub-labels start at the same x0 as the column itself).
        date_col_x0, date_col_x1 = None, None
        for w in all_words:
            if w["text"] in ("Invoice", "Due") and date_col_x0 is None:
                date_col_x0 = w["x0"]
            if w["text"] == "Year" and date_col_x1 is None:
                date_col_x1 = w["x0"]
            if date_col_x0 is not None and date_col_x1 is not None:
                break

        # Group into y-rows with 6 pt tolerance (Crystal Reports may offset
        # words on the same visual line by up to 5 pt)
        row_map: dict = {}
        for w in all_words:
            yk = round(w["top"] / 6) * 6
            row_map.setdefault(yk, []).append(w)

        sorted_ys = sorted(row_map.keys())

        # ── Pass 1: find all APINV rows ──────────────────────────────────────────
        apinv_rows: list = []  # [(yk, invoice_no, voucher_no)]
        for yk in sorted_ys:
            row_ws = row_map[yk]
            # Sort left-to-right so the regex finds "APINV <number>" in order
            row_ws_x = sorted(row_ws, key=lambda w: w["x0"])
            row_tx = " ".join(w["text"] for w in row_ws_x)
            if "APINV" not in row_tx.upper():
                continue

            # Voucher number: 1-4 digit integer immediately before "APINV"
            mv = re.search(r'(?:^|\s)(\d{1,4})\s+APINV\b', row_tx, re.IGNORECASE)
            voucher_no = mv.group(1) if mv else None

            # Regex on joined row text (works when APINV and number share a band)
            m = re.search(r'APINV\s+(\d{7,10})\b', row_tx, re.IGNORECASE)
            if m:
                apinv_rows.append((yk, m.group(1), voucher_no))
                continue

            # Fallback: find "APINV" word, then look right across nearby y-bands
            apinv_x1 = None
            for w in row_ws_x:
                if "APINV" in w["text"].upper():
                    apinv_x1 = w["x1"]
                    break
            if apinv_x1 is not None:
                nearby = [y for y in sorted_ys if abs(y - yk) <= 12]
                found = False
                for ny in nearby:
                    for w in sorted(row_map[ny], key=lambda w: w["x0"]):
                        if w["x0"] >= apinv_x1 - 5 and re.match(r'^\d{7,10}$', w["text"]):
                            apinv_rows.append((yk, w["text"], voucher_no))
                            found = True
                            break
                    if found:
                        break

        # ── Pass 2: extract fields from each bounded block ───────────────────────
        for idx, (yk, invoice_no, voucher_no) in enumerate(apinv_rows):
            next_yk = (apinv_rows[idx + 1][0]
                       if idx + 1 < len(apinv_rows)
                       else max(sorted_ys) + 200)

            block_ys    = [y for y in sorted_ys if yk <= y < next_yk]
            block_words = [w for y in block_ys for w in row_map[y]]
            block_text  = " ".join(w["text"] for w in block_words)

            # Vendor: 10 digits starting with 01
            vendor = None
            mv = re.search(r'\b(01\d{8})\b', block_text)
            if mv:
                vendor = mv.group(1)

            # GL: 6 digits starting with 3
            gl = None
            mg = re.search(r'\b(3\d{5})\b', block_text)
            if mg:
                gl = mg.group(1)

            # CC: known values only
            cc = None
            for mc in re.finditer(r'\b([A-Z]{2}\d{2})\b', block_text):
                if mc.group(1).upper() in known_ccs:
                    cc = mc.group(1).upper()
                    break

            # Invoice date: the date printed ON THE SAME LINE as the APINV /
            # invoice-number row (±6 pt — Crystal Reports row tolerance).
            # Due date: the next date word in the following rows within the
            # block (it sits directly under the invoice date, one line down).
            # Character-based read (handles Party-name overflow into the
            # date column) is tried first; word-based scan is the fallback.
            block_cap = min(yk + 45, next_yk)
            inv_date = _chars_date(yk - 6, yk + 6, date_col_x0, date_col_x1)
            due_date = _chars_date(yk + 6, block_cap, date_col_x0, date_col_x1)

            if inv_date is None:
                same_row_ys = [y for y in sorted_ys if abs(y - yk) <= 6]
                inv_date = _row_date(same_row_ys)
            if due_date is None:
                below_ys = [y for y in sorted_ys if yk + 6 < y <= block_cap]
                due_date = _row_date(below_ys)

            # Total: find the 'Total:' label row (not Sub-Total / Sous-Total)
            total_amt = None
            for y in block_ys:
                row_ws2 = row_map[y]
                row_tx2 = " ".join(w["text"] for w in row_ws2).upper()
                if (re.search(r'\bTOTAL\b', row_tx2)
                        and "SUB" not in row_tx2
                        and "SOUS" not in row_tx2):
                    for w in row_ws2:
                        clean = w["text"].replace(",", "")
                        if re.match(r'^\d+\.\d{2}$', clean):
                            val = float(clean)
                            if val >= 10.0:
                                total_amt = val
                                break
                    if total_amt is not None:
                        break

            # Fallback: largest amount ≥ 10 in the block
            if total_amt is None:
                raw_amounts = re.findall(r'\b([\d,]+\.\d{2})\b', block_text)
                amounts_num = sorted({
                    float(a.replace(",", "")) for a in raw_amounts
                    if float(a.replace(",", "")) >= 10.0
                })
                total_amt = amounts_num[-1] if amounts_num else None

            records[invoice_no] = {
                "voucher":      voucher_no,
                "invoice_date": inv_date,
                "due_date":     due_date,
                "cc":           cc,
                "vendor":       vendor,
                "gl":           gl,
                "total":        total_amt,
            }

    except Exception:
        pass

    return records


def extract_invoice_amounts(pdf_bytes: bytes) -> dict:
    """
    Extract net (subtotal), combined taxes, and total from an Atlantic invoice PDF.
    Returns { "net": str|None, "taxes": float|None, "total": str|None }
    """
    result = {"net": None, "taxes": None, "total": None}
    try:
        with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
            full_text = ""
            for page in pdf.pages:
                full_text += (page.extract_text() or "") + "\n"
        lines = [ln.strip() for ln in full_text.splitlines() if ln.strip()]

        taxes_acc = None
        last_total = None

        for line in lines:
            upper = line.upper()
            amounts = re.findall(r'([\d,]+\.\d{2})', line)
            if not amounts:
                continue
            val = amounts[0].replace(",", "")

            if re.search(r'\bSUB\s*TOTAL\b|\bSOUS.?TOTAL\b', upper):
                result["net"] = val

            elif re.search(r'\bGST\b|\bTPS\b|G\.S\.T|T\.P\.S', upper):
                taxes_acc = (taxes_acc or 0.0) + float(val)

            elif re.search(r'\bQST\b|\bTVQ\b|Q\.S\.T|T\.V\.Q', upper):
                taxes_acc = (taxes_acc or 0.0) + float(val)

            elif re.search(r'^\s*TOTAL\b', upper):
                last_total = val

        if taxes_acc is not None:
            result["taxes"] = round(taxes_acc, 2)
        if last_total:
            result["total"] = last_total

        # Fallback: compute missing value from the other two
        net_f   = float(result["net"])   if result["net"]   else None
        total_f = float(result["total"]) if result["total"] else None
        if net_f is not None and total_f is not None:
            computed_taxes = round(total_f - net_f, 2)
            # Use computed taxes if direct extraction is missing or clearly wrong (>1 CAD off)
            if result["taxes"] is None or abs(result["taxes"] - computed_taxes) > 1.0:
                result["taxes"] = computed_taxes

    except Exception:
        pass
    return result


# ─────────────────────────────────────────────────────────────────────────────
# ── STATEMENT RECONCILIATION FUNCTIONS ────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────

def _recon_parse_amount(raw) -> "float | None":
    """Parse an amount that may be a number already, or text like '1,234.56' /
    '1,234.56-' (Crystal Reports puts the minus sign for credits at the end)."""
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    s = str(raw).strip().replace(",", "")
    if not s:
        return None
    neg = s.endswith("-")
    if neg:
        s = s[:-1]
    try:
        val = float(s)
    except ValueError:
        return None
    return -val if neg else val


_RECON_GRAND_TOTAL_PAT = re.compile(
    r'Total\s+Ow[a-z]*\s*:?\s*(?:CAD)?\s*([\d,]+\.\d{2})', re.IGNORECASE
)


def _recon_find_grand_total(text: str) -> "float | None":
    """Find the statement's own 'Total Owing' figure in free text, so the UI
    can show it next to the sum of the parsed lines as a sanity check that
    the file was read correctly."""
    m = _RECON_GRAND_TOTAL_PAT.search(text)
    if not m:
        return None
    return _recon_parse_amount(m.group(1))


def parse_atlantic_statement_pdf(pdf_bytes: bytes) -> tuple:
    """
    Parse an Atlantic 'Statement of Account' PDF.
    Each line looks like:
      24-03-26 WPD 87059502 R0015751 37,303.43 37,303.43 143
    (date, type code, invoice #, customer reference / PO, invoice amount,
    balance owing, days). Credit-memo lines carry a trailing '-' on the amount.
    Returns (records, grand_total) — grand_total is the 'Total Owing' figure
    printed on the statement itself, or None if it couldn't be found.
    """
    records = []
    grand_total = None
    line_pat = re.compile(
        r'^(\d{2}-\d{2}-\d{2})\s+([A-Z]{3})\s+(\d{6,9})\s+(\S+)\s+'
        r'([\d,]+\.\d{2}-?)\s+([\d,]+\.\d{2}-?)\s+(\d+)\s*$'
    )
    with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            if grand_total is None:
                grand_total = _recon_find_grand_total(text)
            for line in text.splitlines():
                m = line_pat.match(line.strip())
                if not m:
                    continue
                date_str, typ, inv_no, ref, amount_str, balance_str, day = m.groups()
                try:
                    inv_date = datetime.strptime(date_str, "%d-%m-%y").date()
                except ValueError:
                    inv_date = None
                records.append({
                    "invoice_no": inv_no,
                    "type":       typ,
                    "po_ref":     ref,
                    "invoice_date": inv_date,
                    "amount":     _recon_parse_amount(amount_str),
                    "balance":    _recon_parse_amount(balance_str),
                    "day":        int(day),
                })
    return records, grand_total


def parse_atlantic_statement_excel(file_bytes: bytes) -> tuple:
    """
    Parse the Excel export of the same Atlantic statement (B2Win 'b2win'
    sheet). Columns: CAD marker | date | type | invoice # | cust reference |
    invoice amount | balance owing | day.
    Returns (records, grand_total) — see parse_atlantic_statement_pdf.
    """
    records = []
    grand_total = None
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    for sheet in wb.worksheets:
        if "b2win" in sheet.title.lower():
            ws = sheet
            break

    inv_pat = re.compile(r'^\d{6,9}$')
    for row in ws.iter_rows(values_only=True):
        if not row or len(row) < 8:
            continue
        marker, inv_date, typ, inv_no, ref, amount, balance, day = row[:8]
        if str(marker).strip().upper() != "CAD":
            continue
        if grand_total is None and any("Total" in str(c) for c in row if c):
            # Cells split words arbitrarily (e.g. 'Total Ow' | 'ing ') — join
            # with no separator so the word and the trailing amount line up.
            row_text = "".join(str(c) for c in row if c)
            grand_total = _recon_find_grand_total(row_text)
        if not inv_no or not inv_pat.match(str(inv_no).strip()):
            continue
        if hasattr(inv_date, "date"):
            inv_date = inv_date.date()
        records.append({
            "invoice_no": str(inv_no).strip(),
            "type":       str(typ).strip() if typ else None,
            "po_ref":     str(ref).strip() if ref else None,
            "invoice_date": inv_date if isinstance(inv_date, date) else None,
            "amount":     _recon_parse_amount(amount),
            "balance":    _recon_parse_amount(balance),
            "day":        int(day) if day is not None else None,
        })
    return records, grand_total


def parse_atlantic_statement(filename: str, file_bytes: bytes) -> tuple:
    """Dispatch to the PDF or Excel parser based on the file extension.
    Returns (records, grand_total)."""
    if filename.lower().endswith(".pdf"):
        return parse_atlantic_statement_pdf(file_bytes)
    return parse_atlantic_statement_excel(file_bytes)


def _recon_parse_amount_eu(raw) -> "float | None":
    """Parse a European-style amount — space thousands separator, comma
    decimal (Transport Bourret's statement format), e.g. '1 542,81' ->
    1542.81."""
    if raw is None:
        return None
    s = re.sub(r'\s+', '', str(raw)).replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return None


def parse_bourret_statement_pdf(pdf_bytes: bytes) -> tuple:
    """
    Parse a Transport Bourret 'État de compte / Statement' PDF.
    Each line looks like:
      07/21/2026 08/20/2026 14597313 2009575/MDBL236811 1 542,81$
    (invoice date, due date, invoice #, customer reference — often blank,
    sometimes a composite of several tokens — and amount). The reference is
    captured for display only; Bourret's reference is a shipment/BOL number,
    not a PO, so it isn't compared against anything (see check_po in
    RECON_VENDORS).

    Parsed by tokens rather than a single regex: the amount is always the
    last token (ends in ',DD$'); a plain 1-3 digit token right before it is
    a space-grouped thousands prefix ('1 542,81$' -> amount, not part of the
    reference) — everything else between the invoice number and the amount
    is the reference, however many tokens or separators (/, -) it contains.
    A pure regex can't tell a 6-digit numeric reference like '360676' apart
    from a thousands-prefix + amount ('360' + '676 946,01$' both look like
    valid amount shapes) — the token approach resolves it by only ever
    treating a *short* (<=3 digit) trailing token as a thousands group.
    """
    records = []
    grand_total = None
    date_pat        = re.compile(r'^\d{2}/\d{2}/\d{4}$')
    invoice_no_pat  = re.compile(r'^\d{7,9}$')
    amount_end_pat  = re.compile(r'^(\d+),(\d{2})\$$')
    thousands_pat   = re.compile(r'^\d{1,3}$')
    total_pat       = re.compile(r'Total\s+(\d{1,3}(?:\s\d{3})*,\d{2})\$')

    with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            if grand_total is None:
                m = total_pat.search(text)
                if m:
                    grand_total = _recon_parse_amount_eu(m.group(1))
            for line in text.splitlines():
                tokens = line.strip().split()
                if (len(tokens) < 4
                        or not date_pat.match(tokens[0])
                        or not date_pat.match(tokens[1])
                        or not invoice_no_pat.match(tokens[2])):
                    continue
                date_str, inv_no, rest = tokens[0], tokens[2], tokens[3:]
                if not rest or not amount_end_pat.match(rest[-1]):
                    continue
                amount_tokens = [rest[-1]]
                ref_tokens = rest[:-1]
                if ref_tokens and thousands_pat.match(ref_tokens[-1]):
                    amount_tokens.insert(0, ref_tokens.pop())
                try:
                    inv_date = datetime.strptime(date_str, "%m/%d/%Y").date()
                except ValueError:
                    inv_date = None
                records.append({
                    "invoice_no": inv_no,
                    "type":       None,
                    "po_ref":     " ".join(ref_tokens) or None,
                    "invoice_date": inv_date,
                    "amount":     _recon_parse_amount_eu(" ".join(amount_tokens).rstrip("$")),
                    "balance":    None,
                    "day":        None,
                })
    return records, grand_total


def parse_bourret_statement(filename: str, file_bytes: bytes) -> tuple:
    """Dispatch for Transport Bourret's statement. Only PDF has been seen
    from this vendor so far."""
    if filename.lower().endswith(".pdf"):
        return parse_bourret_statement_pdf(file_bytes)
    raise ValueError("Only PDF statements are supported for Transport Bourret right now.")


def parse_proden_statement_pdf(pdf_bytes: bytes) -> tuple:
    """
    Parse a Les Entreprises Proden 'État de compte' aging PDF.
    Each line looks like:
      Apr-20-26 305912 MLQD82865 87 / 117 4,799.06
    (invoice date, invoice #, PO#, a "days overdue / days since invoice"
    pair we don't need, and the outstanding amount). The amount is printed
    in whichever of the four aging-bucket columns applies (Courant / 31-60
    / 61-90 / >90 jours), so exactly one amount value appears per line no
    matter which bucket it's in — we don't need to track which column it
    came from for reconciliation purposes.
    The footer totals row has 5 numbers: a grand total followed by the four
    bucket subtotals; the grand total is used for the statement-total
    sanity check.
    """
    records = []
    grand_total = None
    line_pat = re.compile(
        r'^([A-Za-z]{3}-\d{2}-\d{2})\s+(\d{5,7})\s+(\S+)\s+\d+\s*/\s*\d+\s+([\d,]+\.\d{2})\s*$'
    )
    total_pat = re.compile(
        r'^([\d,]+\.\d{2})\s+[\d,]+\.\d{2}\s+[\d,]+\.\d{2}\s+[\d,]+\.\d{2}\s+[\d,]+\.\d{2}\s*$'
    )
    with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            for line in text.splitlines():
                line = line.strip()
                if grand_total is None:
                    m_tot = total_pat.match(line)
                    if m_tot:
                        grand_total = _recon_parse_amount(m_tot.group(1))
                        continue
                m = line_pat.match(line)
                if not m:
                    continue
                date_str, inv_no, po_ref, amount_str = m.groups()
                try:
                    inv_date = datetime.strptime(date_str, "%b-%d-%y").date()
                except ValueError:
                    inv_date = None
                records.append({
                    "invoice_no": inv_no,
                    "type":       None,
                    "po_ref":     po_ref,
                    "invoice_date": inv_date,
                    "amount":     _recon_parse_amount(amount_str),
                    "balance":    None,
                    "day":        None,
                })
    return records, grand_total


def parse_proden_statement(filename: str, file_bytes: bytes) -> tuple:
    """Dispatch for Proden's statement. Only PDF has been seen from this
    vendor so far."""
    if filename.lower().endswith(".pdf"):
        return parse_proden_statement_pdf(file_bytes)
    raise ValueError("Only PDF statements are supported for Proden right now.")


def parse_distribution_proden_statement_pdf(pdf_bytes: bytes) -> tuple:
    """
    Parse a Distribution Proden Inc. 'État de compte' aging PDF — same
    layout as Les Entreprises Proden's (see parse_proden_statement_pdf),
    with two differences seen in practice:
      - the PO# sometimes has spaces around the hyphen (e.g. 'ML - 320643'
        instead of 'ML-320437'), so it's captured as everything between the
        invoice # and the aging-days fraction rather than a single
        whitespace-free token, then has its internal spaces stripped;
      - invoice numbers are zero-padded to 6 digits on the statement
        ('005201') but stored unpadded in the accounting system ('5201'),
        so leading zeros are stripped for the lookup to match.
    Distribution Proden splits its statement by branch/customer account, so
    a reconciliation is normally built from several of these PDFs (one per
    branch) parsed and combined by the caller before matching against a
    single system extract.
    """
    records = []
    grand_total = None
    line_pat = re.compile(
        r'^([A-Za-z]{3}-\d{2}-\d{2})\s+(\d{5,7})\s+(.+?)\s+\d+\s*/\s*\d+\s+([\d,]+\.\d{2})\s*$'
    )
    total_pat = re.compile(
        r'^([\d,]+\.\d{2})\s+[\d,]+\.\d{2}\s+[\d,]+\.\d{2}\s+[\d,]+\.\d{2}\s+[\d,]+\.\d{2}\s*$'
    )
    with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            for line in text.splitlines():
                line = line.strip()
                if grand_total is None:
                    m_tot = total_pat.match(line)
                    if m_tot:
                        grand_total = _recon_parse_amount(m_tot.group(1))
                        continue
                m = line_pat.match(line)
                if not m:
                    continue
                date_str, inv_no, po_ref, amount_str = m.groups()
                try:
                    inv_date = datetime.strptime(date_str, "%b-%d-%y").date()
                except ValueError:
                    inv_date = None
                records.append({
                    "invoice_no": str(int(inv_no)),
                    "type":       None,
                    "po_ref":     re.sub(r'\s+', '', po_ref),
                    "invoice_date": inv_date,
                    "amount":     _recon_parse_amount(amount_str),
                    "balance":    None,
                    "day":        None,
                })
    return records, grand_total


def parse_distribution_proden_statement(filename: str, file_bytes: bytes) -> tuple:
    """Dispatch for Distribution Proden's statement. Only PDF has been seen
    from this vendor so far."""
    if filename.lower().endswith(".pdf"):
        return parse_distribution_proden_statement_pdf(file_bytes)
    raise ValueError("Only PDF statements are supported for Distribution Proden right now.")


def parse_pyrogaz_statement_pdf(pdf_bytes: bytes) -> tuple:
    """
    Parse a Pyrogaz Inc. 'État de compte' aging PDF.
    Each line looks like:
      29/05/2026 023878 1442.52
    (invoice date, invoice #, outstanding amount — printed in whichever of
    the four aging-bucket columns applies: Courant / 30 jours / 60 jours /
    90 jours, so exactly one amount appears per line). There is no PO /
    customer-reference column at all (see check_po in RECON_VENDORS).
    The footer prints the four bucket subtotals on one line, followed by
    the grand total alone on the next line; the grand total is used for
    the statement-total sanity check.
    Invoice numbers are zero-padded to 6 digits on the statement ('023878')
    but stored unpadded in the accounting system ('23878'), same as
    Distribution Proden — leading zeros are stripped for the lookup to match.
    """
    records = []
    grand_total = None
    line_pat = re.compile(r'^(\d{2}/\d{2}/\d{4})\s+(\d{5,7})\s+([\d,]+\.\d{2})\s*$')
    subtotal_pat = re.compile(
        r'^[\d,]+\.\d{2}\s+[\d,]+\.\d{2}\s+[\d,]+\.\d{2}\s+[\d,]+\.\d{2}\s*$'
    )
    grand_total_pat = re.compile(r'^([\d,]+\.\d{2})\s*$')
    with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            after_subtotal = False
            for line in text.splitlines():
                line = line.strip()
                if not line:
                    continue
                if subtotal_pat.match(line):
                    after_subtotal = True
                    continue
                if after_subtotal:
                    m_tot = grand_total_pat.match(line)
                    if m_tot:
                        grand_total = _recon_parse_amount(m_tot.group(1))
                    after_subtotal = False
                    continue
                m = line_pat.match(line)
                if not m:
                    continue
                date_str, inv_no, amount_str = m.groups()
                try:
                    inv_date = datetime.strptime(date_str, "%d/%m/%Y").date()
                except ValueError:
                    inv_date = None
                records.append({
                    "invoice_no": str(int(inv_no)),
                    "type":       None,
                    "po_ref":     None,
                    "invoice_date": inv_date,
                    "amount":     _recon_parse_amount(amount_str),
                    "balance":    None,
                    "day":        None,
                })
    return records, grand_total


def parse_pyrogaz_statement(filename: str, file_bytes: bytes) -> tuple:
    """Dispatch for Pyrogaz's statement. Only PDF has been seen from this
    vendor so far."""
    if filename.lower().endswith(".pdf"):
        return parse_pyrogaz_statement_pdf(file_bytes)
    raise ValueError("Only PDF statements are supported for Pyrogaz right now.")


def parse_system_extract(file_bytes: bytes) -> dict:
    """
    Parse the accounting-system AP extract (any vendor — the export is the
    same GL/voucher layout regardless of who the invoices belong to).
    Returns { invoice_no: {"total", "costctr", "paid", "payment_date"} }.
    An invoice can span several GL allocation lines (same invoice #, same
    total, different account/cost-centre) — those are collapsed into one
    record; a payment on ANY of its lines marks the invoice as paid.
    """
    df = pd.read_excel(BytesIO(file_bytes))
    cols = {c.strip().lower(): c for c in df.columns}

    def col(*names):
        for n in names:
            if n in cols:
                return cols[n]
        return None

    c_inv   = col("invoice")
    c_total = col("invoicetotal")
    c_cc    = col("costctr")
    c_pay   = col("paymentno")
    c_paydt = col("paymentdate")
    if c_inv is None or c_total is None:
        return {}

    records = {}
    for _, row in df.iterrows():
        inv_no = str(row[c_inv]).strip() if pd.notna(row[c_inv]) else ""
        if not inv_no:
            continue
        total = float(row[c_total]) if pd.notna(row[c_total]) else 0.0
        paid  = c_pay is not None and pd.notna(row[c_pay])
        rec = records.get(inv_no)
        if rec is None:
            rec = {
                "invoice_no":   inv_no,
                "total":        total,
                "costctr":      row[c_cc] if c_cc is not None and pd.notna(row[c_cc]) else None,
                "paid":         False,
                "payment_date": None,
            }
            records[inv_no] = rec
        if paid:
            rec["paid"] = True
            if c_paydt is not None and pd.notna(row[c_paydt]):
                rec["payment_date"] = row[c_paydt]
    return records


def parse_extraction_excel(file_bytes: bytes) -> dict:
    """
    Parse the invoice-copy tracking Excel (folder listing exported from the
    Atlantic invoices share). Returns { invoice_no: {"po", "location", "filename"} }.
    Expects columns named 'Name' (file name), 'Colonne2' (PO extracted from
    the file name) and 'Colonne4' (folder location relative to the share root).
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    header = [str(h).strip() if h else "" for h in rows[0]]

    def idx(*names):
        for n in names:
            if n in header:
                return header.index(n)
        return None

    idx_name = idx("Name")
    idx_po   = idx("Colonne2")
    idx_loc  = idx("Colonne4")
    if idx_name is None:
        return {}

    inv_pat = re.compile(r'^(\d{6,9})\b')
    records = {}
    for row in rows[1:]:
        name = row[idx_name] if idx_name < len(row) else None
        if not name:
            continue
        m = inv_pat.match(str(name).strip())
        if not m:
            continue
        inv_no = m.group(1)
        po  = row[idx_po]  if idx_po  is not None and idx_po  < len(row) else None
        loc = row[idx_loc] if idx_loc is not None and idx_loc < len(row) else None
        records[inv_no] = {
            "po":       str(po).strip() if po else None,
            "location": str(loc).strip() if loc else None,
            "filename": str(name).strip(),
        }
    return records


def reconcile_statement(statement: list, system: dict, extraction: dict, check_po: bool = True) -> dict:
    """
    Drive the reconciliation off the vendor statement (per the workflow: the
    statement is what the vendor sends us, so every line in it must be
    accounted for). For each statement line:
      1. Look it up in the copy-tracking extract by invoice number (this is
         independent of registration status — every entry gets a has_copy /
         extraction_po / location, even when the copy-tracking file wasn't
         supplied at all, in which case has_copy is simply False for all).
      2. Look it up in the system extract by invoice number.
         - Not found            -> pending_registration (has_copy tells you
           whether the file is on hand and just needs entering, or missing
           entirely — no separate top-level bucket for that distinction).
         - Found + amount matches -> matched
         - Found + amount differs -> amount_mismatch
      3. When check_po is True and there's a copy on file, compare its PO
         (from the file name) against the statement's PO/customer-reference
         -> po_mismatch (this can happen on a matched invoice too, not just
         a pending one). Some vendors' "reference" isn't a PO at all (e.g.
         Transport Bourret's is a shipment/BOL number) — check_po=False
         skips this comparison entirely for those.
    """
    buckets = {
        "matched": [], "amount_mismatch": [],
        "pending_registration": [], "po_mismatch": [],
    }
    for line in statement:
        inv_no = line["invoice_no"]
        entry = dict(line)

        ext_rec = extraction.get(inv_no)
        entry["has_copy"]      = ext_rec is not None
        entry["extraction_po"] = ext_rec["po"] if ext_rec else None
        entry["location"]      = ext_rec["location"] if ext_rec else None
        entry["filename"]      = ext_rec["filename"] if ext_rec else None

        if check_po:
            po_a = str(line.get("po_ref") or "").strip().upper()
            po_b = str((ext_rec or {}).get("po") or "").strip().upper()
            entry["po_mismatch"] = bool(ext_rec is not None and po_a and po_b and po_a != po_b)
        else:
            entry["po_mismatch"] = False

        sys_rec = system.get(inv_no)
        if sys_rec is not None:
            entry["registered"]   = True
            entry["system_total"] = sys_rec["total"]
            entry["costctr"]      = sys_rec["costctr"]
            entry["paid"]         = sys_rec["paid"]
            entry["payment_date"] = sys_rec["payment_date"]
            if abs(abs(line["amount"] or 0) - abs(sys_rec["total"] or 0)) > 0.02:
                buckets["amount_mismatch"].append(entry)
            else:
                buckets["matched"].append(entry)
        else:
            entry["registered"] = False
            buckets["pending_registration"].append(entry)

        if entry["po_mismatch"]:
            buckets["po_mismatch"].append(entry)

    return buckets


def build_recon_summary_rows(buckets: dict, statement_total: float, check_po: bool = True) -> list:
    """
    Build the reconciliation summary as a flat, hierarchical list of
    (label, count, amount) rows — mirroring the AP team's own manual
    reconciliation layout: a top-level Total / Reconciled / Not Reconciled
    split with a balancing check, then a breakdown of Not Reconciled by
    whether we have a copy on file, then a 'To Review' section for the
    cross-cutting discrepancy flags. Blank label = section break; None in
    count/amount = leave that cell empty. The PO Mismatch line is omitted
    when check_po is False (vendors whose "reference" isn't a PO).
    """
    matched   = buckets["matched"]
    mismatch  = buckets["amount_mismatch"]
    pending   = buckets["pending_registration"]
    po_mism   = buckets["po_mismatch"]
    pending_with_copy = [r for r in pending if r["has_copy"]]
    pending_no_copy   = [r for r in pending if not r["has_copy"]]

    def amt(rows):
        return round(sum((r.get("amount") or 0) for r in rows), 2)

    mismatch_amt = amt(mismatch)
    # "Reconciled" money-wise means "accounted for in the system" — so its
    # amount includes invoices registered with a discrepancy too (those are
    # still in the system, just flagged in "To Review" for the exact figure).
    # Its count, however, is the clean-match count only, so the discrepancy
    # count is visible on its own line in "To Review" instead of hiding
    # inside "Reconciled". Not Reconciled is simpler: not in the system at
    # all, so count and amount both mean the same population there.
    reconciled_amt      = round(amt(matched) + mismatch_amt, 2)
    not_reconciled_amt  = amt(pending)
    check = round(statement_total - reconciled_amt - not_reconciled_amt, 2)
    check = 0.0 if abs(check) < 0.005 else check

    rows = [
        ("Total Statement",                    len(matched) + len(pending) + len(mismatch), round(statement_total, 2)),
        ("Reconciled",                         len(matched), reconciled_amt),
        ("Not Reconciled",                     len(pending), not_reconciled_amt),
        ("Reconciliation Check",               None, check),
        ("", None, None),
        ("Not Reconciled — breakdown",         None, None),
        ("  Pending Registration (has copy)",  len(pending_with_copy), amt(pending_with_copy)),
        ("  No Copy on File",                  len(pending_no_copy), amt(pending_no_copy)),
        ("", None, None),
        ("To Review",                          None, None),
        ("  Amount Mismatch",                  len(mismatch), mismatch_amt),
    ]
    if check_po:
        rows.append(("  PO Mismatch", len(po_mism), amt(po_mism)))
    return rows


def make_recon_report(buckets: dict, statement_total: float, grand_total: "float | None",
                       check_po: bool = True) -> bytes:
    """Build the downloadable Excel report: a Summary sheet (hierarchical
    counts/totals plus the statement-total sanity check) plus one sheet per
    category. The PO Mismatch sheet is omitted when check_po is False."""
    sheet_specs = [
        ("Matched",              buckets["matched"]),
        ("Amount Mismatch",      buckets["amount_mismatch"]),
        ("Pending Registration", buckets["pending_registration"]),
    ]
    if check_po:
        sheet_specs.append(("PO Mismatch", buckets["po_mismatch"]))
    cols = [
        "invoice_no", "type", "invoice_date", "po_ref", "amount", "day",
        "registered", "system_total", "costctr", "paid", "payment_date",
        "has_copy", "extraction_po", "location", "po_mismatch", "filename",
    ]
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        summary_rows = [
            {"Category": label, "Invoice Count": count, "Total Amount (CAD)": amount}
            for label, count, amount in build_recon_summary_rows(buckets, statement_total, check_po)
        ]
        summary_rows.append({"Category": "", "Invoice Count": None, "Total Amount (CAD)": None})
        summary_rows.append({
            "Category": "Statement Total (per file)" if grand_total is not None
                        else "Statement Total (not found in file)",
            "Invoice Count": None,
            "Total Amount (CAD)": grand_total,
        })
        summary_rows.append({
            "Category": "Sum of Parsed Invoice Lines",
            "Invoice Count": None,
            "Total Amount (CAD)": round(statement_total, 2),
        })
        pd.DataFrame(summary_rows).to_excel(writer, sheet_name="Summary", index=False)
        for label, rows in sheet_specs:
            if rows:
                df = pd.DataFrame(rows)
                df = df[[c for c in cols if c in df.columns]]
            else:
                df = pd.DataFrame(columns=cols)
            df.to_excel(writer, sheet_name=label[:31], index=False)
    buf.seek(0)
    return buf.read()


# Registry of vendors the Reconciliation module can parse a statement for.
# The system extract and the copy-tracking Excel are vendor-agnostic (the
# user exports them already scoped to one vendor), only the statement needs
# a vendor-specific parser — add new vendors here as their templates are built.
# check_po: whether the statement's reference column is a PO worth comparing
# against the copy-tracking file name (False for vendors like Transport
# Bourret, whose reference is a shipment/BOL number, not a PO).
RECON_VENDORS = {
    "atlantic": {"label": "Atlantic Packaging",  "parse_statement": parse_atlantic_statement, "check_po": True},
    "bourret":  {"label": "Transport Bourret",   "parse_statement": parse_bourret_statement,  "check_po": False},
    "proden":   {"label": "Les Entreprises Proden", "parse_statement": parse_proden_statement, "check_po": True},
    "distribution_proden": {"label": "Distribution Proden", "parse_statement": parse_distribution_proden_statement, "check_po": True},
    "pyrogaz":  {"label": "Pyrogaz Inc.",       "parse_statement": parse_pyrogaz_statement,  "check_po": False},
}


# ─────────────────────────────────────────────────────────────────────────────
# CONTROL FACTURAS — OTROS PROVEEDORES (non-Atlantic vendors)
# Reception / coding / vendor-control tool, replacing a fragile shared .xlsm
# that relied on VBA, a live link into a 105k-row external workbook, and a
# Power Query folder scan over a network share.
# ─────────────────────────────────────────────────────────────────────────────
CF_KNOWN_LEGACY_SHEETS = {"CONTROL", "FOURNISSEUR", "BASE", "VALIDATION"}


def _cf_next_id() -> int:
    nid = st.session_state.cf_next_id
    st.session_state.cf_next_id += 1
    return nid


def _cf_build_nom_pdf(fournisseur, nro_facture, division, po_reception) -> str:
    """
    Auto-generate the invoice's PDF filename from its own fields — nombre
    corto del proveedor, número de factura, división y PO — instead of it
    being typed by hand. Mirrors the naming convention already used for
    invoice copies (e.g. 'UBA 000028314 EV MD552802.pdf').
    """
    parts = [str(p).strip() for p in (fournisseur, nro_facture, division, po_reception) if str(p or "").strip()]
    return (" ".join(parts) + ".pdf") if parts else ""


def _cf_proveedor_lookup(codigo: str):
    codigo = str(codigo or "").strip().upper()
    for p in st.session_state.cf_proveedores:
        if str(p.get("codigo", "")).strip().upper() == codigo:
            return p
    return None


def _cf_fmt_date(v) -> str:
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return str(v).strip() if v not in (None, "") else ""


def _cf_import_legacy_excel(file_bytes: bytes) -> dict:
    """
    One-time migration of the user's existing CONTROL_FACTURES workbook
    (CONTROL, FOURNISSEUR and any extra per-vendor rule sheet such as
    'Xerox') into this app's own storage. Read-only — nothing is written
    back to the uploaded file.
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    result = {"facturas": 0, "proveedores": 0, "reglas": 0, "responsables": 0, "estados": 0, "warnings": []}

    if "FOURNISSEUR" in wb.sheetnames:
        ws = wb["FOURNISSEUR"]
        new_prov = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            codigo = row[0] if len(row) > 0 else None
            nombre = row[1] if len(row) > 1 else None
            # Some FOURNISSEUR sheets carry a small legend area (e.g. cells
            # literally containing "Fournisseur" / "# Vendor" as labels,
            # not data) above or beside the real vendor rows — a genuine
            # vendor always has both a code and a name, so require both.
            if codigo and str(codigo).strip() and nombre and str(nombre).strip():
                new_prov.append({
                    "codigo":     str(codigo).strip(),
                    "nombre":     str(nombre).strip(),
                    "vendor_no":  str(row[2]).strip() if len(row) > 2 and row[2] not in (None, "") else "",
                    "divisiones": "",
                    "activo":     True,
                    "notas":      "",
                })
        if new_prov:
            st.session_state.cf_proveedores = new_prov
            result["proveedores"] = len(new_prov)

    if "CONTROL" in wb.sheetnames:
        ws = wb["CONTROL"]
        header_row = None
        for r in range(1, min(ws.max_row, 30) + 1):
            vals = [c.value for c in ws[r]]
            if "NRO FACTURE" in vals and "FOURNISSEUR" in vals:
                header_row = r
                break
        if header_row:
            headers = [str(c.value).strip() if c.value else "" for c in ws[header_row]]
            col_idx = {h: i for i, h in enumerate(headers) if h}

            def gv(row_vals, key, default=None):
                i = col_idx.get(key)
                return row_vals[i] if i is not None and i < len(row_vals) else default

            new_facturas = []
            seen_responsables, seen_estados = set(), set()
            for row_vals in ws.iter_rows(min_row=header_row + 1, values_only=True):
                fournisseur = gv(row_vals, "FOURNISSEUR")
                nro_facture = gv(row_vals, "NRO FACTURE")
                if not fournisseur and not nro_facture:
                    continue
                prix = gv(row_vals, "Prix A/T")
                responsable = str(gv(row_vals, "Responsable") or "").strip()
                etat = str(gv(row_vals, "ETAT") or "").strip()
                if responsable:
                    seen_responsables.add(responsable)
                if etat:
                    seen_estados.add(etat)
                division = str(gv(row_vals, "DIVISION") or "").strip()
                po_reception = str(gv(row_vals, "PO - Reception") or "").strip()
                new_facturas.append({
                    # A full import REPLACES the list, so the id is just the
                    # row's position — matching the Excel's own correlativo
                    # (column A, "=MAX($A12:A$13)+1") instead of an
                    # ever-growing counter left over from earlier imports.
                    "id":                  len(new_facturas) + 1,
                    "fournisseur":         str(fournisseur or "").strip(),
                    "nom_system":          str(gv(row_vals, "NOM SYSTEM") or "").strip(),
                    "nro_vendor":          str(gv(row_vals, "NRO FOURNISSEUR") or "").strip(),
                    "nro_facture":         str(nro_facture or "").strip(),
                    "date_facture":        _cf_fmt_date(gv(row_vals, "DATE FACTURE")),
                    "division":            division,
                    "po_reception":        po_reception,
                    "responsable":         responsable,
                    "etat":                etat,
                    "problema":            False,
                    "comentario":          str(gv(row_vals, "Coment") or "").strip(),
                    "date_reception":      _cf_fmt_date(gv(row_vals, "DATE RECEPTION")),
                    "ultima_actualizacion": _cf_fmt_date(gv(row_vals, "Dernier Misa a jour")),
                    "poste":               bool(gv(row_vals, "POSTÉ")),
                    "payee":               False,  # legacy column was #REF! on every row — starts clean
                    "nom_pdf":             _cf_build_nom_pdf(fournisseur, nro_facture, division, po_reception),
                    "cc":                  str(gv(row_vals, "CC") or "").strip(),
                    "gl":                  str(gv(row_vals, "GL") or "").strip(),
                    "monto":               prix if isinstance(prix, (int, float)) else None,
                })
            if new_facturas:
                st.session_state.cf_facturas = new_facturas
                st.session_state.cf_next_id = len(new_facturas) + 1
                result["facturas"] = len(new_facturas)
            if seen_responsables:
                existing = set(st.session_state.cf_responsables)
                st.session_state.cf_responsables = sorted(existing | seen_responsables)
                result["responsables"] = len(seen_responsables - existing)
            if seen_estados:
                existing = set(st.session_state.cf_estados)
                st.session_state.cf_estados = st.session_state.cf_estados + sorted(seen_estados - existing)
                result["estados"] = len(seen_estados - existing)
        else:
            result["warnings"].append("No se encontró la fila de encabezados (FOURNISSEUR / NRO FACTURE) en la hoja CONTROL.")

    for sheet_name in wb.sheetnames:
        if sheet_name in CF_KNOWN_LEGACY_SHEETS:
            continue
        ws = wb[sheet_name]
        headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
        if "GL" not in headers or "CC" not in headers:
            continue
        col_idx = {h: i for i, h in enumerate(headers) if h}

        def g(row_vals, key):
            i = col_idx.get(key)
            return row_vals[i] if i is not None and i < len(row_vals) else None

        new_reglas = []
        last_concept = ""
        for row_vals in ws.iter_rows(min_row=2, values_only=True):
            concept = g(row_vals, "Concept")
            if concept:
                last_concept = str(concept).strip()
            gl_val, cc_val = g(row_vals, "GL"), g(row_vals, "CC")
            if not gl_val and not cc_val:
                continue
            no_cliente = g(row_vals, "No. Client")
            new_reglas.append({
                "proveedor":  sheet_name,
                "concepto":   last_concept,
                "gl":         str(gl_val).strip() if gl_val not in (None, "") else "",
                "cc":         str(cc_val).strip() if cc_val not in (None, "") else "",
                "no_cliente": str(no_cliente).strip() if no_cliente not in (None, "") else "",
                "notas":      "",
            })
        if new_reglas:
            st.session_state.cf_reglas.extend(new_reglas)
            result["reglas"] += len(new_reglas)

    return result


def _cf_create_stamp_pdf(lines: list, x: float, y: float, w: float, h: float, page_w: float, page_h: float) -> bytes:
    """
    Build a single-page PDF holding just the red-bordered coding stamp
    (same look as the red box in E1:F7 of the legacy CONTROL sheet, and
    the same visual style as the Atlantic coding stamp) at an arbitrary,
    caller-chosen position — since non-Atlantic invoices don't share a
    common layout, the position is picked by hand per invoice rather than
    fixed in Settings.
    """
    packet = BytesIO()
    c = canvas.Canvas(packet, pagesize=(page_w, page_h))
    y_bottom = page_h - y - h  # y is measured from the TOP, reportlab draws from the bottom
    c.setStrokeColorRGB(0.85, 0.0, 0.0)
    c.setFillColorRGB(1.0, 1.0, 1.0)
    c.setLineWidth(1.8)
    c.rect(x, y_bottom, w, h, fill=1)
    c.setFillColorRGB(0.85, 0.0, 0.0)
    c.setFont("Helvetica-Bold", 8.5)
    pad = 8
    line_h = (h - pad) / max(len(lines), 1)
    tx, ty = x + pad, y_bottom + h - line_h
    for i, line in enumerate(lines):
        c.drawString(tx, ty - i * line_h, line)
    c.save()
    packet.seek(0)
    return packet.read()


def _cf_render_stamp_preview(pdf_bytes: bytes, x: float, y: float, w: float, h: float, dpi: int = 110) -> Image.Image:
    """
    Render the invoice's first page as an image with a translucent red box
    overlaid where the stamp would land, so the position can be picked by
    eye before committing — each vendor's invoice has its own layout, so
    there's no fixed spot to default to. x/y/w/h are in PDF points
    (x, y measured from the page's top-left corner).
    """
    images = convert_from_bytes(pdf_bytes, first_page=1, last_page=1, dpi=dpi)
    img = images[0].convert("RGBA")
    scale = dpi / 72
    overlay = Image.new("RGBA", img.size, (0, 0, 0, 0))
    draw = ImageDraw.Draw(overlay)
    box = [x * scale, y * scale, (x + w) * scale, (y + h) * scale]
    draw.rectangle(box, outline=(217, 0, 0, 255), width=3, fill=(217, 0, 0, 70))
    return Image.alpha_composite(img, overlay).convert("RGB")


def _cf_cross_reference_ap(file_bytes: bytes, sheet_name: str, vendor_col: str, ref_col: str, match_col: str):
    """
    Fill in the PO/Reception of pending cf_facturas rows by matching
    (vendor + reference) against an uploaded Purchases (AP) export.
    Only the matched values are kept in cf_facturas — the uploaded
    workbook itself is discarded by the caller once this returns.
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb[sheet_name]
    header_cells = next(ws.iter_rows(min_row=1, max_row=1))
    headers = [str(c.value).strip() if c.value else "" for c in header_cells]
    idx = {h: i for i, h in enumerate(headers)}
    vi, ri, mi = idx.get(vendor_col), idx.get(ref_col), idx.get(match_col)
    if vi is None or ri is None or mi is None:
        return 0, ["No se encontraron las columnas indicadas en la hoja seleccionada."]

    lookup = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) <= max(vi, ri, mi):
            continue
        key = f"{row[vi]} {row[ri]}".strip().upper()
        if key and key not in lookup and row[mi] not in (None, ""):
            lookup[key] = row[mi]

    matched = 0
    for f in st.session_state.cf_facturas:
        if f.get("po_reception"):
            continue
        key = f"{f.get('fournisseur', '')} {f.get('nro_facture', '')}".strip().upper()
        if key in lookup:
            f["po_reception"] = str(lookup[key])
            matched += 1
    return matched, []


# ─────────────────────────────────────────────────────────────────────────────
# CSS STYLES
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
.stTabs [data-baseweb="tab"] { font-size: 15px; font-weight: 600; padding: 8px 18px; }
.inv-card {
    background: #f8f9fa; border-radius: 10px; padding: 14px 18px;
    margin-bottom: 8px; border-left: 5px solid #ccc;
}
.inv-ok   { border-left-color: #28a745 !important; }
.inv-warn { border-left-color: #ffc107 !important; }
.inv-err  { border-left-color: #dc3545 !important; }
.stamp-preview {
    border: 2px solid #cc0000; padding: 8px 12px;
    display: inline-block; background: white;
    font-family: monospace; font-weight: bold; color: #cc0000;
    font-size: 13px; line-height: 1.6; border-radius: 3px;
}
.match-row {
    background: #f0fff4; border-radius: 8px; padding: 10px 14px;
    margin-bottom: 6px; border-left: 4px solid #28a745;
    font-size: 14px;
}
.pending-row {
    background: #fff8f0; border-radius: 8px; padding: 10px 14px;
    margin-bottom: 6px; border-left: 4px solid #ffc107;
    font-size: 14px;
}
.stat-box {
    background: white; border-radius: 10px; padding: 16px 20px;
    text-align: center; box-shadow: 0 1px 4px rgba(0,0,0,0.08);
    border: 1px solid #e9ecef;
}
.stat-num  { font-size: 36px; font-weight: 700; line-height: 1.1; }
.stat-lbl  { font-size: 13px; color: #6c757d; margin-top: 2px; }
.green { color: #28a745; }
.amber { color: #e08000; }
.red   { color: #dc3545; }
.blue  { color: #0d6efd; }
.split-row {
    background: #f8f9fa; border-radius: 8px; padding: 10px 14px;
    margin-bottom: 6px; border-left: 4px solid #0d6efd;
    font-size: 14px;
}
.split-warn {
    border-left-color: #ffc107 !important;
    background: #fffdf0 !important;
}
.module-card [data-testid="stVerticalBlockBorderWrapper"] {
    transition: box-shadow 0.15s ease, transform 0.15s ease;
}
.module-card:hover [data-testid="stVerticalBlockBorderWrapper"] {
    box-shadow: 0 4px 14px rgba(0,0,0,0.12);
    transform: translateY(-2px);
}
.module-icon { font-size: 34px; line-height: 1; }
.module-label { font-size: 17px; font-weight: 700; margin-top: 6px; }
.module-desc  { font-size: 13px; color: #6c757d; margin-top: 2px; min-height: 34px; }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# MODULE REGISTRY
# ─────────────────────────────────────────────────────────────────────────────
MODULES = [
    {"key": "splitter", "icon": "✂️",  "label": "Invoice Splitter",  "desc": "Split batch PDFs into individual invoice files."},
    {"key": "matcher",  "icon": "🔗", "label": "Invoice Matcher",   "desc": "Match invoices with POs and merge into one PDF."},
    {"key": "coding",   "icon": "🏷️",  "label": "Invoice Coding",    "desc": "Stamp GL / Cost Centre codes on invoices."},
    {"key": "couru",    "icon": "📊", "label": "Couru Code",        "desc": "Extract coding data for Couru entry."},
    {"key": "audit",    "icon": "🔍", "label": "AP Audit",          "desc": "Validate the A/P voucher audit listing."},
    {"key": "recon",    "icon": "🧮", "label": "Statement Reconciliation", "desc": "Reconcile Atlantic's statement against the system and invoice copies."},
    {"key": "payment",  "icon": "💳", "label": "Payment Packager",  "desc": "Bundle invoices into payment batches."},
    {"key": "database", "icon": "🗄️",  "label": "Database",          "desc": "Manage vendors, GL codes & users."},
    {"key": "settings", "icon": "⚙️",  "label": "Settings",          "desc": "Configure the coding stamp position."},
    {"key": "control_prov", "icon": "📋", "label": "Control Facturas — Otros Proveedores",
     "desc": "Recepción, codificación y control de proveedores distintos a Atlantic."},
]

# ─────────────────────────────────────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## 📂 Modules")
    nav_labels = ["🏠 Home"] + [f"{m['icon']} {m['label']}" for m in MODULES]
    nav_keys   = [None] + [m["key"] for m in MODULES]
    nav_idx    = nav_keys.index(st.session_state.active_module) \
        if st.session_state.active_module in nav_keys else 0
    picked = st.selectbox("Go to", nav_labels, index=nav_idx, label_visibility="collapsed")
    picked_key = nav_keys[nav_labels.index(picked)]
    if picked_key != st.session_state.active_module:
        st.session_state.active_module = picked_key
        st.rerun()

    st.divider()
    st.markdown("## ⚙️ Work Session")

    user_list = st.session_state.usuarios + ["✏️ Other..."]
    sel_idx = st.selectbox(
        "👤 Posted By",
        range(len(user_list)),
        format_func=lambda x: user_list[x],
    )
    if user_list[sel_idx] == "✏️ Other...":
        current_user = st.text_input("Name:", placeholder="Enter name")
    else:
        current_user = user_list[sel_idx]

    coding_date = st.date_input("📅 Coding Date", value=date.today())

    st.divider()

    n_proc = len(st.session_state.processed)
    st.metric("Coded Invoices", n_proc)

    if n_proc > 0:
        zip_bytes = _get_processed_zip()
        st.download_button(
            "⬇️ Download ZIP (all)",
            data=zip_bytes,
            file_name=f"invoices_{date.today().strftime('%Y%m%d')}.zip",
            mime="application/zip",
            use_container_width=True,
        )
        if st.button("🗑️ Clear all results", use_container_width=True, type="secondary"):
            st.session_state.processed = []
            st.rerun()

    st.divider()

    with st.expander("💾 Save / Load Database"):
        db_export = {
            "proveedores":     st.session_state.proveedores,
            "gl_codes":        st.session_state.gl_codes,
            "usuarios":        st.session_state.usuarios,
            "cf_facturas":     st.session_state.cf_facturas,
            "cf_proveedores":  st.session_state.cf_proveedores,
            "cf_reglas":       st.session_state.cf_reglas,
            "cf_responsables": st.session_state.cf_responsables,
            "cf_estados":      st.session_state.cf_estados,
        }
        st.download_button(
            "⬇️ Export DB (JSON)",
            data=json.dumps(db_export, indent=2, ensure_ascii=False),
            file_name="invoice_db.json",
            mime="application/json",
            use_container_width=True,
        )
        db_upload = st.file_uploader("📥 Import DB (JSON)", type=["json"], key="db_import")
        if db_upload:
            try:
                db = json.loads(db_upload.read())
                if "proveedores"    in db: st.session_state.proveedores    = db["proveedores"]
                if "gl_codes"       in db: st.session_state.gl_codes       = db["gl_codes"]
                if "usuarios"       in db: st.session_state.usuarios       = db["usuarios"]
                if "cf_facturas"     in db: st.session_state.cf_facturas     = db["cf_facturas"]
                if "cf_proveedores"  in db: st.session_state.cf_proveedores  = db["cf_proveedores"]
                if "cf_reglas"       in db: st.session_state.cf_reglas       = db["cf_reglas"]
                if "cf_responsables" in db: st.session_state.cf_responsables = db["cf_responsables"]
                if "cf_estados"      in db: st.session_state.cf_estados      = db["cf_estados"]
                if db.get("cf_facturas"):
                    st.session_state.cf_next_id = max(
                        (r.get("id", 0) for r in db["cf_facturas"]), default=0
                    ) + 1
                st.success("✅ Database loaded")
            except Exception as e:
                st.error(f"Error: {e}")

# ─────────────────────────────────────────────────────────────────────────────
# MAIN HEADER
# ─────────────────────────────────────────────────────────────────────────────
active_module = st.session_state.active_module

if active_module is None:
    st.markdown("""
    <h1 style='margin-bottom:0'>📄 Atlantic — Invoice Tools</h1>
    <p style='color:gray;margin-top:4px'>Pick a module to get started</p>
    """, unsafe_allow_html=True)

    n_cols = 4
    rows = [MODULES[i:i + n_cols] for i in range(0, len(MODULES), n_cols)]
    for row in rows:
        cols = st.columns(n_cols)
        for col, mod in zip(cols, row):
            with col:
                st.markdown('<div class="module-card">', unsafe_allow_html=True)
                with st.container(border=True):
                    st.markdown(
                        f"<div class='module-icon'>{mod['icon']}</div>"
                        f"<div class='module-label'>{mod['label']}</div>"
                        f"<div class='module-desc'>{mod['desc']}</div>",
                        unsafe_allow_html=True,
                    )
                    if st.button("Open →", key=f"open_{mod['key']}", use_container_width=True):
                        st.session_state.active_module = mod["key"]
                        st.rerun()
                st.markdown('</div>', unsafe_allow_html=True)
else:
    active_mod = next(m for m in MODULES if m["key"] == active_module)
    col_home, col_title = st.columns([1, 6])
    with col_home:
        if st.button("🏠 Home", use_container_width=True):
            st.session_state.active_module = None
            st.rerun()
    with col_title:
        st.markdown(
            f"<h2 style='margin:2px 0 0 0'>{active_mod['icon']} {active_mod['label']}</h2>",
            unsafe_allow_html=True,
        )
    st.divider()

# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 — INVOICE SPLITTER
# ══════════════════════════════════════════════════════════════════════════════
if active_module == "splitter":
    st.subheader("✂️ Split Batch Invoice PDFs")
    st.markdown(
        "Upload one or more **batch PDFs** from Atlantic (each may contain multiple invoices). "
        "The tool will detect each invoice automatically, split them into individual PDFs, "
        "and name each file as **`{Invoice No} {CC} {Bill of Lading}.pdf`** — "
        "ready to use directly in the Matcher tab."
    )

    col_up, col_clear = st.columns([5, 1])
    with col_up:
        split_uploads = st.file_uploader(
            "Drag or select one or more batch PDFs",
            type=["pdf"],
            accept_multiple_files=True,
            key=f"split_uploader_{st.session_state.splitter_upload_key}",
        )
    with col_clear:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🗑️ Clear", use_container_width=True, key="split_clear"):
            st.session_state.splitter_results  = []
            st.session_state.splitter_zip      = None
            st.session_state.splitter_upload_key += 1
            st.rerun()

    if split_uploads:
        # Scanned batches have no text layer — detect that up front (once
        # per upload set) and, if found, offer a rotation preview/selector.
        # Scanners often output pages sideways; getting this wrong would
        # silently OCR (and save) every invoice upside down, so it's
        # confirmed visually rather than fully automatic.
        upload_sig = tuple((f.name, f.size) for f in split_uploads)
        if st.session_state.get("_splitter_preview_sig") != upload_sig:
            st.session_state["_splitter_preview_sig"] = upload_sig
            scanned_bytes = None
            for f in split_uploads:
                fb = f.getvalue()
                with pdfplumber.open(BytesIO(fb)) as pdf:
                    sample_text = "".join((p.extract_text() or "") for p in pdf.pages[:3])
                if len(sample_text.strip()) < 30:
                    scanned_bytes = fb
                    break
            st.session_state["_splitter_scanned_sample"] = scanned_bytes
            st.session_state["splitter_rotation"] = (
                _splitter_suggest_rotation(scanned_bytes)
                if scanned_bytes and OCR_AVAILABLE else 0
            )

        scanned_sample = st.session_state.get("_splitter_scanned_sample")
        rotation = 0
        if scanned_sample and OCR_AVAILABLE:
            st.info(
                "📷 Detected a scanned PDF (no digital text) — OCR will be used to read it. "
                "**Click the thumbnail that looks upright** before splitting (applies to all "
                "scanned files in this batch). The auto-suggested one is only a guess — "
                "OCR can sometimes misread text even at the wrong angle, so please verify visually."
            )
            base_preview = _splitter_render_page_image(scanned_sample, 1, dpi=100)
            rot_options = [0, 90, 180, 270]
            if base_preview is not None:
                thumb_cols = st.columns(4)
                for col, angle in zip(thumb_cols, rot_options):
                    with col:
                        thumb = base_preview.rotate(angle, expand=True) if angle else base_preview
                        st.image(thumb, caption=f"{angle}°" if angle else "0° (no change)",
                                  use_container_width=True)
            rotation = st.radio(
                "Rotation to apply",
                options=rot_options,
                index=rot_options.index(st.session_state.get("splitter_rotation", 0)),
                format_func=lambda a: f"{a}°" if a else "0° (no change)",
                horizontal=True,
                key="splitter_rotation_radio",
            )
            st.session_state["splitter_rotation"] = rotation
        elif scanned_sample and not OCR_AVAILABLE:
            st.warning(
                "⚠️ This PDF looks scanned (no digital text) but OCR isn't available in "
                "this environment — invoices won't be detected."
            )

        do_split = st.button("✂️ Split Invoices", type="primary", use_container_width=False)

        if do_split:
            st.session_state.splitter_results = []
            st.session_state.splitter_zip     = None
            all_results = []
            prog = st.progress(0, text="Processing…")

            for f_idx, f in enumerate(split_uploads):
                prog.progress((f_idx) / len(split_uploads),
                              text=f"Splitting {f.name}…")

                def _split_progress(cur_page, total_pg, _fname=f.name, _fi=f_idx):
                    frac = (_fi + (cur_page / total_pg if total_pg else 1)) / len(split_uploads)
                    prog.progress(min(frac, 0.999),
                                  text=f"OCR {_fname}: page {cur_page}/{total_pg}…")

                try:
                    batch_results = split_batch_pdf(
                        f.read(), rotation=rotation, progress_callback=_split_progress
                    )
                    for r in batch_results:
                        r["source_file"] = f.name
                    all_results.extend(batch_results)
                except Exception as e:
                    st.error(f"Error processing **{f.name}**: {e}")

            prog.progress(1.0, text="✅ Done")
            st.session_state.splitter_results = all_results
            if all_results:
                st.session_state.splitter_zip = make_splitter_zip(all_results)
            st.rerun()

    # ── Results display ───────────────────────────────────────────────────────
    results = st.session_state.splitter_results

    if results:
        n_ok   = sum(1 for r in results if not r["warning"])
        n_warn = sum(1 for r in results if r["warning"])

        # Summary stats
        c1, c2, c3 = st.columns(3)
        with c1:
            st.markdown(
                f"<div class='stat-box'>"
                f"<div class='stat-num blue'>{len(results)}</div>"
                f"<div class='stat-lbl'>Invoices detected</div></div>",
                unsafe_allow_html=True,
            )
        with c2:
            st.markdown(
                f"<div class='stat-box'>"
                f"<div class='stat-num green'>{n_ok}</div>"
                f"<div class='stat-lbl'>Ready to download</div></div>",
                unsafe_allow_html=True,
            )
        with c3:
            st.markdown(
                f"<div class='stat-box'>"
                f"<div class='stat-num amber'>{n_warn}</div>"
                f"<div class='stat-lbl'>Need review</div></div>",
                unsafe_allow_html=True,
            )

        st.markdown("<br>", unsafe_allow_html=True)

        # Download all ZIP
        if st.session_state.splitter_zip:
            st.download_button(
                f"⬇️ Download all as ZIP ({len(results)} invoices)",
                data=st.session_state.splitter_zip,
                file_name=f"split_invoices_{date.today().strftime('%Y%m%d')}.zip",
                mime="application/zip",
                type="primary",
            )

        st.divider()

        # Per-invoice rows
        for row_idx, item in enumerate(results):
            row_cls = "split-row split-warn" if item["warning"] else "split-row"
            icon    = "⚠️" if item["warning"] else "✅"

            col_info, col_dl = st.columns([5, 1])
            with col_info:
                pages_str = ", ".join(str(p) for p in item["source_pages"])
                page_lbl  = f"{item['page_count']} page{'s' if item['page_count'] > 1 else ''}"
                warn_html = (
                    f"<br><span style='color:#e08000;font-size:12px'>{item['warning']}</span>"
                    if item["warning"] else ""
                )
                st.markdown(
                    f"<div class='{row_cls}'>"
                    f"{icon} <b>{item['filename']}</b>"
                    f"&nbsp;&nbsp;|&nbsp;&nbsp;"
                    f"Invoice: <code>{item['invoice_no']}</code> &nbsp; "
                    f"CC: <code>{item['cc']}</code> &nbsp; "
                    f"BOL: <code>{item['bol']}</code> &nbsp; "
                    f"· {page_lbl} (source p. {pages_str})"
                    f"{warn_html}"
                    f"</div>",
                    unsafe_allow_html=True,
                )
            with col_dl:
                st.download_button(
                    "⬇️",
                    data=item["pdf_bytes"],
                    file_name=item["filename"],
                    mime="application/pdf",
                    key=f"spdl_{row_idx}",
                    use_container_width=True,
                    help=f"Download {item['filename']}",
                )

    elif not split_uploads:
        st.info("📂 Upload one or more Atlantic batch PDFs to get started.")


# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 — INVOICE MATCHER
# ══════════════════════════════════════════════════════════════════════════════
if active_module == "matcher":
    st.subheader("🔗 Invoice & PO Matcher")
    st.markdown(
        "Upload your invoice PDFs and your PO (Purchase Order) PDFs. "
        "The tool will match them by PO number, flatten both documents to ensure "
        "consistent rendering, merge them into a single PDF, and package everything for download."
    )

    if not PIKEPDF_AVAILABLE:
        st.warning(
            "⚠️ **pikepdf** is not installed. Flattening will use rasterization (pdf2image) as fallback. "
            "Add `pikepdf` to `requirements.txt` for best results."
        )

    st.info(
        "📋 **Expected invoice filename format:** `{Invoice No} {Cost Centre} {PO Number}.pdf`  \n"
        "Example: `82196530 ML V0020978.pdf`  →  will search for PO file `V0020978.pdf`"
    )

    col_inv, col_po = st.columns(2)
    with col_inv:
        st.markdown("**📄 Invoice PDFs**")
        inv_files = st.file_uploader(
            "Upload invoices",
            type=["pdf"],
            accept_multiple_files=True,
            key=f"match_inv_{st.session_state.matcher_upload_key}",
            label_visibility="collapsed",
        )
        if inv_files:
            st.caption(f"{len(inv_files)} invoice(s) loaded")

    with col_po:
        st.markdown("**📦 PO PDFs**")
        po_files = st.file_uploader(
            "Upload POs",
            type=["pdf"],
            accept_multiple_files=True,
            key=f"match_po_{st.session_state.matcher_upload_key}",
            label_visibility="collapsed",
        )
        if po_files:
            st.caption(f"{len(po_files)} PO(s) loaded")

    col_run, col_clr, _ = st.columns([2, 2, 5])
    with col_run:
        run_match = st.button(
            "🔗 Run Matching",
            type="primary",
            use_container_width=True,
            disabled=not bool(inv_files),
        )
    with col_clr:
        if st.button("🗑️ Clear results", use_container_width=True):
            st.session_state.matcher_results    = None
            st.session_state.matcher_zip        = None
            st.session_state.matcher_upload_key += 1
            st.rerun()

    if run_match and inv_files:
        if not po_files:
            st.warning("⚠️ No PO files uploaded — all invoices will be marked as pending.")

        prog_bar  = st.progress(0, text="Matching…")
        prog_text = st.empty()

        def _progress(cur, tot, fname):
            prog_bar.progress((cur + 1) / tot, text=f"Processing {fname}…")
            prog_text.caption(f"({cur + 1}/{tot})")

        results = run_matching(inv_files, po_files or [], progress_callback=_progress)
        prog_bar.progress(1.0, text="✅ Done")
        prog_text.empty()

        st.session_state.matcher_results = results
        st.session_state.matcher_zip     = make_matcher_zip(results)
        st.rerun()

    # ── Results ────────────────────────────────────────────────────────────────
    if st.session_state.matcher_results:
        results     = st.session_state.matcher_results
        matched     = results["matched"]
        pending     = results["pending"]
        unmatched_po = results["unmatched_po"]

        # Stats
        c1, c2, c3 = st.columns(3)
        with c1:
            st.markdown(
                f"<div class='stat-box'><div class='stat-num green'>{len(matched)}</div>"
                f"<div class='stat-lbl'>Matched</div></div>",
                unsafe_allow_html=True,
            )
        with c2:
            st.markdown(
                f"<div class='stat-box'><div class='stat-num amber'>{len(pending)}</div>"
                f"<div class='stat-lbl'>Unmatched invoices</div></div>",
                unsafe_allow_html=True,
            )
        with c3:
            st.markdown(
                f"<div class='stat-box'><div class='stat-num blue'>{len(unmatched_po)}</div>"
                f"<div class='stat-lbl'>Unused POs</div></div>",
                unsafe_allow_html=True,
            )

        st.markdown("<br>", unsafe_allow_html=True)

        if st.session_state.matcher_zip:
            st.download_button(
                "⬇️ Download all results (ZIP)",
                data=st.session_state.matcher_zip,
                file_name=f"matched_invoices_{date.today().strftime('%Y%m%d')}.zip",
                mime="application/zip",
                type="primary",
            )

        st.divider()

        if matched:
            with st.expander(f"✅ Matched ({len(matched)})", expanded=True):
                for item in matched:
                    col1, col2 = st.columns([5, 1])
                    with col1:
                        st.markdown(
                            f"<div class='match-row'>"
                            f"✅ <b>{item['invoice_name']}</b>"
                            f"&nbsp;&nbsp;+&nbsp;&nbsp;"
                            f"📦 <code>{item['po_name']}</code>"
                            f"</div>",
                            unsafe_allow_html=True,
                        )
                    with col2:
                        st.download_button(
                            "⬇️",
                            data=item["merged_bytes"],
                            file_name=item["invoice_name"],
                            mime="application/pdf",
                            key=f"mdl_{item['invoice_name']}",
                            use_container_width=True,
                        )

        if pending:
            with st.expander(f"⚠️ Unmatched Invoices ({len(pending)})", expanded=True):
                st.caption("These invoices had no corresponding PO file.")
                for item in pending:
                    st.markdown(
                        f"<div class='pending-row'>"
                        f"📄 <b>{item['invoice_name']}</b>"
                        f"&nbsp;&nbsp;|&nbsp;&nbsp;"
                        f"PO searched: <code>{item['po_id']}</code>"
                        f"&nbsp;&nbsp;—&nbsp;&nbsp;"
                        f"{item['reason']}"
                        f"</div>",
                        unsafe_allow_html=True,
                    )

        if unmatched_po:
            with st.expander(f"📦 Unused POs ({len(unmatched_po)})"):
                st.caption("These PO files were uploaded but no invoice referenced them.")
                for po_name in unmatched_po:
                    st.markdown(f"- `{po_name}.pdf`")


# ══════════════════════════════════════════════════════════════════════════════
# TAB 3 — INVOICE CODING (upload + review + results unified)
# ══════════════════════════════════════════════════════════════════════════════
if active_module == "coding":
    if "upload_key" not in st.session_state:
        st.session_state.upload_key = 0

    # ── Top toolbar: upload + clear all ──────────────────────────────────────
    col_up, col_clear_upload, col_clear_all = st.columns([5, 1, 1])
    with col_up:
        uploaded = st.file_uploader(
            "Drag or select one or more PDFs",
            type=["pdf"],
            accept_multiple_files=True,
            help="Supports bulk upload. Only the first page of each invoice is stamped.",
            key=f"uploader_{st.session_state.upload_key}",
        )
    with col_clear_upload:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🗑️ Clear\nfiles", use_container_width=True,
                     help="Remove uploaded files to load a new batch"):
            st.session_state.upload_key += 1
            st.rerun()
    with col_clear_all:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🗑️ Clear\nall", use_container_width=True,
                     help="Clear uploaded files AND all coded results"):
            st.session_state.processed    = []
            st.session_state.upload_key  += 1
            st.rerun()

    # ── Upload / review section ───────────────────────────────────────────────
    if not uploaded:
        st.session_state.pop("_inv_ui_cache", None)
        st.session_state.pop("_inv_ui_sig", None)
        if not st.session_state.processed:
            st.info("📂 Upload invoices to get started. You can select multiple files at once.")
    else:
        st.success(f"✅ **{len(uploaded)} file(s)** loaded — analyzing…")
        st.divider()

        # Re-parse PDFs only when files OR the GL/vendor database changes
        _sig = (
            tuple((f.name, f.size) for f in uploaded),
            tuple(r["codigo"] for r in st.session_state.gl_codes),
            tuple(r["prefijo"] for r in st.session_state.proveedores),
        )
        if _sig != st.session_state.get("_inv_ui_sig"):
            _inv_ui = []
            for f in uploaded:
                raw = f.read()
                data = extract_invoice_data(raw, f.name)
                data["filename"] = f.name
                data["raw_bytes"] = raw
                is_six = data.get("is_six", False)
                if is_six:
                    data["vendor_auto"] = VENDOR_EXCEPCION
                    prefix = data.get("cc_prefix")
                    _, c = get_vendor_cc(prefix) if prefix else (None, None)
                    data["cc_auto"]  = c
                    data["needs_cc"] = (c is None)
                else:
                    prefix = data.get("cc_prefix")
                    v, c = get_vendor_cc(prefix) if prefix else (None, None)
                    data["vendor_auto"] = v
                    data["cc_auto"]     = c
                    data["needs_cc"]    = (v is None or c is None)
                data["gl_auto"] = get_gl(data.get("product_code"))
                _inv_ui.append(data)
            st.session_state["_inv_ui_cache"] = _inv_ui
            st.session_state["_inv_ui_sig"]   = _sig
        invoices_ui = st.session_state["_inv_ui_cache"]

        resolved_cc     = {}
        resolved_vendor = {}
        resolved_gl     = {}

        n_auto   = sum(
            1 for inv in invoices_ui
            if not inv.get("error")
            and (inv.get("invoice_no") or inv.get("customer_order"))
            and not inv["needs_cc"]
            and inv.get("gl_auto")
        )
        n_review = len(invoices_ui) - n_auto

        if n_auto:
            st.success(f"✅ {n_auto} invoice(s) auto-coded — no review needed.")
        if n_review:
            st.warning(f"⚠️ {n_review} invoice(s) require manual input — review below before coding.")

        for idx, inv in enumerate(invoices_ui):
            has_err    = inv.get("error") or (not inv.get("invoice_no") and not inv.get("customer_order"))
            needs_input = inv["needs_cc"] or not inv.get("gl_auto")
            auto_coded  = not has_err and not needs_input

            if auto_coded:
                # Pre-populate resolved dicts; no UI needed
                resolved_vendor[idx] = VENDOR_EXCEPCION if inv.get("is_six") else inv["vendor_auto"]
                resolved_cc[idx]     = inv["cc_auto"]
                resolved_gl[idx]     = inv["gl_auto"]
                continue

            icon = "❌" if has_err else "⚠️"
            with st.expander(f"{icon}  {inv['filename']}", expanded=True):
                c1, c2, c3 = st.columns([1.2, 1.2, 1])

                with c1:
                    st.markdown("**📑 Extracted from PDF**")
                    st.write(f"Invoice No: `{inv.get('invoice_no') or '—'}`")
                    st.write(f"Customer Order: `{inv.get('customer_order') or '—'}`")
                    st.write(f"CC Prefix: `{inv.get('cc_prefix') or '—'}`")
                    st.write(f"Product Code: `{inv.get('product_code') or '—'}`")
                    if inv.get("is_six"):
                        st.info("⚡ Type-6 invoice — vendor exception")
                    if inv.get("error"):
                        st.error(f"Error: {inv['error']}")
                    if inv.get("ocr_used"):
                        st.info("🔍 Text extracted via OCR (scanned PDF)")

                with c2:
                    st.markdown("**🏷️ Vendor / Cost Centre**")
                    if inv["is_six"]:
                        st.write(f"Vendor (fixed): `{VENDOR_EXCEPCION}`")
                        resolved_vendor[idx] = VENDOR_EXCEPCION
                        if inv["cc_auto"]:
                            st.success(f"CC: `{inv['cc_auto']}`")
                            resolved_cc[idx] = inv["cc_auto"]
                        else:
                            st.warning("CC prefix not detected — select manually")
                            cc_opts = sorted(set(r["cc"] for r in st.session_state.proveedores))
                            sel_cc = st.selectbox("Select CC:", cc_opts, key=f"cc6_{idx}")
                            resolved_cc[idx] = sel_cc
                    elif inv["vendor_auto"] and inv["cc_auto"]:
                        st.success(f"Vendor: `{inv['vendor_auto']}`")
                        st.success(f"CC: `{inv['cc_auto']}`")
                        resolved_vendor[idx] = inv["vendor_auto"]
                        resolved_cc[idx]     = inv["cc_auto"]
                    else:
                        st.warning(f"Prefix `{inv.get('cc_prefix')}` not found")
                        all_opts = sorted(set(r["cc"] for r in st.session_state.proveedores))
                        sel_cc = st.selectbox("Manual CC:", all_opts, key=f"ccman_{idx}")
                        sel_vendor = next(
                            (r["vendor"] for r in st.session_state.proveedores if r["cc"] == sel_cc),
                            "UNKNOWN"
                        )
                        resolved_cc[idx]     = sel_cc
                        resolved_vendor[idx] = sel_vendor

                with c3:
                    st.markdown("**📊 GL Account**")
                    if inv.get("gl_auto"):
                        st.success(f"GL: `{inv['gl_auto']}`")
                        resolved_gl[idx] = inv["gl_auto"]
                    else:
                        st.warning("GL not detected automatically")
                        gl_opts = sorted(set(r["gl"] for r in st.session_state.gl_codes))
                        sel_gl = st.selectbox("Manual GL:", gl_opts, key=f"glman_{idx}")
                        resolved_gl[idx] = sel_gl

                cc_prev   = resolved_cc.get(idx, "??")
                gl_prev   = resolved_gl.get(idx, "??")
                vd_prev   = resolved_vendor.get(idx, "??")
                usr_prev  = current_user or "???"
                date_prev = coding_date.strftime("%d/%m/%Y")
                st.markdown(f"""
                <div style='margin-top:10px'>
                <p style='margin-bottom:4px; color:gray; font-size:12px'>👁️ Stamp preview:</p>
                <div class='stamp-preview'>
                POSTED BY: {usr_prev}<br>
                VENDOR: {vd_prev}<br>
                CC: {cc_prev}&nbsp;&nbsp;|&nbsp;&nbsp;GL: {gl_prev}<br>
                DATE: {date_prev}
                </div></div>
                """, unsafe_allow_html=True)

        st.divider()
        col_btn, col_info = st.columns([1, 3])
        with col_btn:
            do_process = st.button(
                "🚀 Code Invoices",
                type="primary",
                use_container_width=True,
                disabled=not bool(current_user),
            )
        with col_info:
            if not current_user:
                st.warning("⚠️ Select a responsible user (Posted By) in the sidebar before processing.")

        if do_process and current_user:
            progress = st.progress(0, text="Starting…")
            errors = []
            for idx, inv in enumerate(invoices_ui):
                fname  = inv["filename"]
                progress.progress((idx + 1) / len(invoices_ui), text=f"Coding {fname}…")
                cc     = resolved_cc.get(idx, "???")
                vendor = resolved_vendor.get(idx, "???")
                gl     = resolved_gl.get(idx, "???")
                try:
                    stamped = process_one(inv["raw_bytes"], current_user, vendor, cc, gl, coding_date)
                    st.session_state.processed.append({
                        "filename":       fname,
                        "original_bytes": inv["raw_bytes"],
                        "pdf_bytes":      stamped,
                        "invoice_no":     inv.get("invoice_no"),
                        "vendor":         vendor,
                        "cc":             cc,
                        "gl":             gl,
                        "user":           current_user,
                        "date":           coding_date.strftime("%d/%m/%Y"),
                        "date_obj":       coding_date,
                        "ts":             datetime.now().strftime("%Y-%m-%d %H:%M"),
                    })
                except Exception as e:
                    errors.append(f"{fname}: {e}")
            progress.progress(1.0, text="✅ Done")
            if errors:
                for err in errors:
                    st.error(err)
            else:
                st.success(f"🎉 **{len(invoices_ui)} invoice(s)** coded successfully.")
                st.balloons()

    # ── Results section (always visible when there are coded invoices) ────────
    if st.session_state.processed:
        n = len(st.session_state.processed)
        st.divider()
        col_hdr, col_zip, col_del = st.columns([3, 2, 1])
        with col_hdr:
            st.subheader(f"📋 Coded Invoices — {n} file(s)")
        with col_zip:
            zip_all = _get_processed_zip()
            st.download_button(
                f"⬇️ Download ZIP ({n})",
                data=zip_all,
                file_name=f"coded_invoices_{date.today().strftime('%Y%m%d')}.zip",
                mime="application/zip",
                type="primary",
                use_container_width=True,
            )
        with col_del:
            if st.button("🗑️ Delete all", use_container_width=True):
                st.session_state.processed = []
                st.rerun()

        with st.expander("📅 Modify Coding Date"):
            st.caption("Use this if the invoice could not be registered on the same day it was coded.")
            col_nd, col_sel = st.columns([1, 2])
            with col_nd:
                new_date = st.date_input("New date:", value=date.today(), key="new_date_picker")
            with col_sel:
                inv_labels = [f"{item['filename']}  ({item['date']})" for item in st.session_state.processed]
                sel_labels = st.multiselect("Select invoices to update:", inv_labels)
            if st.button("🔄 Regenerate with new date", type="primary"):
                count = 0
                for i, item in enumerate(st.session_state.processed):
                    label = f"{item['filename']}  ({item['date']})"
                    if label in sel_labels:
                        try:
                            new_stamped = process_one(
                                item["original_bytes"], item["user"],
                                item["vendor"], item["cc"], item["gl"], new_date,
                            )
                            st.session_state.processed[i]["pdf_bytes"] = new_stamped
                            st.session_state.processed[i]["date"]      = new_date.strftime("%d/%m/%Y")
                            st.session_state.processed[i]["date_obj"]  = new_date
                            count += 1
                        except Exception as e:
                            st.error(f"Error in {item['filename']}: {e}")
                if count:
                    st.success(f"✅ {count} invoice(s) regenerated with date {new_date.strftime('%d/%m/%Y')}")
                    st.rerun()

        st.divider()
        to_delete = []
        for i, item in enumerate(st.session_state.processed):
            col1, col2, col3, col4 = st.columns([4, 1.5, 1, 1])
            with col1:
                st.markdown(f"📄 **{item['filename']}**")
                st.caption(
                    f"Vendor: `{item['vendor']}` | CC: `{item['cc']}` | "
                    f"GL: `{item['gl']}` | By: **{item['user']}** | "
                    f"Date: `{item['date']}` | Coded: {item['ts']}"
                )
            with col2:
                st.download_button(
                    "⬇️ Download",
                    data=item["pdf_bytes"],
                    file_name=item["filename"],
                    mime="application/pdf",
                    key=f"dl_{i}",
                    use_container_width=True,
                )
            with col3:
                if st.button("🗑️", key=f"del_{i}", help="Remove from list"):
                    to_delete.append(i)
            with col4:
                st.caption(f"#{i+1}")
            st.markdown("<hr style='margin:6px 0; border-color:#eee'>", unsafe_allow_html=True)

        if to_delete:
            for i in sorted(to_delete, reverse=True):
                st.session_state.processed.pop(i)
            st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# TAB 4 — COURU CODE
# ══════════════════════════════════════════════════════════════════════════════
if active_module == "couru":
    st.subheader("📊 Couru Code — Invoice Report")
    st.markdown(
        "Upload invoice PDFs to extract key data and download an Excel report "
        "with invoice number, date, GL account, cost centre, and subtotal before taxes."
    )

    if "couru_upload_key" not in st.session_state:
        st.session_state.couru_upload_key = 0

    col_up, col_clear = st.columns([5, 1])
    with col_up:
        couru_files = st.file_uploader(
            "Drag or select one or more invoice PDFs",
            type=["pdf"],
            accept_multiple_files=True,
            key=f"couru_uploader_{st.session_state.couru_upload_key}",
        )
    with col_clear:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🗑️ Clear", use_container_width=True, key="couru_clear"):
            st.session_state.couru_upload_key += 1
            st.session_state.pop("couru_results", None)
            st.rerun()

    if not couru_files:
        st.info("📂 Upload invoice PDFs to get started.")
    else:
        st.success(f"✅ **{len(couru_files)} file(s)** loaded")

        if st.button("📊 Generate Excel Report", type="primary", key="couru_run"):
            couru_data = []
            prog = st.progress(0, text="Extracting…")
            for fi, f in enumerate(couru_files):
                prog.progress((fi + 1) / len(couru_files), text=f"Processing {f.name}…")
                raw = f.read()
                d = extract_couru_data(raw, f.name)
                couru_data.append({
                    "Invoice No": d["invoice_no"] or "—",
                    "Date":       d["date"]       or "—",
                    "GL":         d["gl"]         or "—",
                    "CC":         d["cc"]         or "—",
                    "Subtotal":   d["subtotal"]   or "—",
                    "Error":      d["error"]      or "",
                })
            prog.progress(1.0, text="✅ Done")
            st.session_state["couru_results"] = couru_data

        if st.session_state.get("couru_results"):
            rows = st.session_state["couru_results"]
            df = pd.DataFrame(rows)

            n_ok   = sum(1 for r in rows if not r["Error"] and r["Invoice No"] != "—")
            n_warn = sum(1 for r in rows if r["Error"] or r["GL"] == "—" or r["CC"] == "—")

            c1, c2, c3 = st.columns(3)
            with c1:
                st.markdown(
                    f"<div class='stat-box'><div class='stat-num blue'>{len(rows)}</div>"
                    f"<div class='stat-lbl'>Invoices processed</div></div>",
                    unsafe_allow_html=True,
                )
            with c2:
                st.markdown(
                    f"<div class='stat-box'><div class='stat-num green'>{n_ok}</div>"
                    f"<div class='stat-lbl'>Complete</div></div>",
                    unsafe_allow_html=True,
                )
            with c3:
                st.markdown(
                    f"<div class='stat-box'><div class='stat-num amber'>{n_warn}</div>"
                    f"<div class='stat-lbl'>Need review</div></div>",
                    unsafe_allow_html=True,
                )

            st.markdown("<br>", unsafe_allow_html=True)
            st.dataframe(df, use_container_width=True, hide_index=True)

            buf = BytesIO()
            display_df = df.drop(columns=["Error"]) if all(r == "" for r in df["Error"]) else df
            with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                display_df.to_excel(writer, index=False)
            buf.seek(0)
            st.download_button(
                f"⬇️ Download Excel ({len(rows)} invoices)",
                data=buf.read(),
                file_name=f"couru_report_{date.today().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
            )


# ══════════════════════════════════════════════════════════════════════════════
# TAB 5 — AP AUDIT VALIDATION
# ══════════════════════════════════════════════════════════════════════════════
if active_module == "audit":
    st.subheader("🔍 AP Audit Validation")
    st.markdown(
        "Cross-reference invoice PDFs against the **A/P Voucher Audit Listing** "
        "to verify invoice number, date, payment date, cost centre, vendor, and GL "
        "before the session is permanently posted."
    )

    col_inv_up, col_rpt_up = st.columns(2)
    with col_inv_up:
        st.markdown("**📄 Invoice PDFs**")
        audit_inv_files = st.file_uploader(
            "Upload invoices",
            type=["pdf"],
            accept_multiple_files=True,
            key="audit_inv_upload",
            label_visibility="collapsed",
        )
        if audit_inv_files:
            st.caption(f"{len(audit_inv_files)} invoice(s) loaded")

    with col_rpt_up:
        st.markdown("**📋 A/P Voucher Audit Listing (PDF)**")
        audit_report_file = st.file_uploader(
            "Upload audit report",
            type=["pdf"],
            accept_multiple_files=False,
            key="audit_report_upload",
            label_visibility="collapsed",
        )
        if audit_report_file:
            st.caption(f"Report: **{audit_report_file.name}**")

    col_run_a, col_clr_a, _ = st.columns([2, 2, 5])
    with col_run_a:
        run_audit = st.button(
            "🔍 Run Validation",
            type="primary",
            use_container_width=True,
            disabled=not (audit_inv_files and audit_report_file),
        )
    with col_clr_a:
        if st.button("🗑️ Clear results", use_container_width=True, key="audit_clear"):
            st.session_state["audit_results"]    = None
            st.session_state["audit_data_count"] = 0
            st.rerun()

    if run_audit and audit_inv_files and audit_report_file:
        audit_data = parse_audit_report(audit_report_file.read())
        st.session_state["audit_data_count"] = len(audit_data)

        if not audit_data:
            st.warning(
                "⚠️ No records could be parsed from the audit report PDF. "
                "Please verify the file is the correct A/P Voucher Audit Listing."
            )
        else:
            def _cmp(a, b):
                if a is None or b is None:
                    return None
                if isinstance(a, date) and isinstance(b, date):
                    return a == b
                return str(a).strip().upper() == str(b).strip().upper()

            def _cmp_amt(a, b, tol=0.05):
                """Compare two floats with tolerance; None on either side → None."""
                if a is None or b is None:
                    return None
                try:
                    return abs(float(a) - float(b)) <= tol
                except (TypeError, ValueError):
                    return None

            val_results = []
            prog = st.progress(0, text="Validating invoices…")
            for fi, f in enumerate(audit_inv_files):
                prog.progress((fi + 1) / len(audit_inv_files), text=f"Processing {f.name}…")
                raw = f.read()

                inv_data   = extract_invoice_data(raw, f.name)
                invoice_no = inv_data.get("invoice_no")
                inv_date   = extract_invoice_date(raw)
                exp_due    = calc_due_date(inv_date)

                # Amounts from invoice PDF
                amts = extract_invoice_amounts(raw)
                inv_net   = float(amts["net"])   if amts["net"]   else None
                inv_taxes = amts["taxes"]  # already float or None
                inv_total = float(amts["total"]) if amts["total"] else None
                # If total not found, compute it
                if inv_total is None and inv_net is not None and inv_taxes is not None:
                    inv_total = round(inv_net + inv_taxes, 2)

                cc_prefix = inv_data.get("cc_prefix")
                if inv_data.get("is_six"):
                    vendor_ext = VENDOR_EXCEPCION
                    _, cc_ext  = get_vendor_cc(cc_prefix) if cc_prefix else (None, None)
                else:
                    vendor_ext, cc_ext = get_vendor_cc(cc_prefix) if cc_prefix else (None, None)
                gl_ext = get_gl(inv_data.get("product_code"))

                audit_rec    = audit_data.get(invoice_no) if invoice_no else None
                aud          = audit_rec or {}
                voucher_aud  = aud.get("voucher")
                inv_date_aud = aud.get("invoice_date")
                due_date_aud = aud.get("due_date")
                cc_aud       = aud.get("cc")
                vendor_aud   = aud.get("vendor")
                gl_aud       = aud.get("gl")
                aud_total    = aud.get("total")  # float or None (largest amount in audit row)

                total_ok = _cmp_amt(inv_total, aud_total)

                val_results.append({
                    "filename":     f.name,
                    "invoice_no":   invoice_no or "—",
                    "found":        audit_rec is not None,
                    # From invoice PDF
                    "inv_date":     inv_date,
                    "exp_due":      exp_due,
                    "cc_ext":       cc_ext,
                    "vendor_ext":   vendor_ext,
                    "gl_ext":       gl_ext,
                    "inv_net":      inv_net,
                    "inv_taxes":    inv_taxes,
                    "inv_total":    inv_total,
                    # From audit report
                    "voucher_aud":  voucher_aud,
                    "inv_date_aud": inv_date_aud,
                    "due_date_aud": due_date_aud,
                    "cc_aud":       cc_aud,
                    "vendor_aud":   vendor_aud,
                    "gl_aud":       gl_aud,
                    "aud_total":    aud_total,
                    # Validation flags
                    "inv_date_ok":  _cmp(inv_date, inv_date_aud),
                    "due_date_ok":  _cmp(exp_due, due_date_aud),
                    "cc_ok":        _cmp(cc_ext, cc_aud),
                    "vendor_ok":    _cmp(vendor_ext, vendor_aud),
                    "gl_ok":        _cmp(gl_ext, gl_aud),
                    "total_ok":     total_ok,
                })

            prog.progress(1.0, text="✅ Validation complete")
            st.session_state["audit_results"] = val_results
            st.rerun()

    # ── Results ────────────────────────────────────────────────────────────────
    if st.session_state.get("audit_results"):
        val_results  = st.session_state["audit_results"]
        audit_count  = st.session_state.get("audit_data_count", 0)

        def _icon(ok):
            if ok is None: return "❓"
            return "✅" if ok else "❌"

        def _fmt_date(d):
            return d.strftime("%d/%m/%Y") if d else "—"

        def _fmt_amt(v):
            if v is None: return "—"
            return f"{float(v):,.2f}"

        _all_flags = ["inv_date_ok", "due_date_ok", "cc_ok", "vendor_ok", "gl_ok", "total_ok"]

        n_found  = sum(1 for r in val_results if r["found"])
        n_all_ok = sum(
            1 for r in val_results
            if r["found"] and all(v is not False for v in [r[k] for k in _all_flags])
        )
        n_issues = len(val_results) - n_all_ok

        st.info(f"Audit report: **{audit_count}** record(s) parsed")
        st.markdown("<br>", unsafe_allow_html=True)

        c1, c2, c3, c4 = st.columns(4)
        with c1:
            st.markdown(
                f"<div class='stat-box'><div class='stat-num blue'>{len(val_results)}</div>"
                f"<div class='stat-lbl'>Invoices checked</div></div>",
                unsafe_allow_html=True,
            )
        with c2:
            col_c = "green" if n_found == len(val_results) else "amber"
            st.markdown(
                f"<div class='stat-box'><div class='stat-num {col_c}'>{n_found}</div>"
                f"<div class='stat-lbl'>Found in audit</div></div>",
                unsafe_allow_html=True,
            )
        with c3:
            st.markdown(
                f"<div class='stat-box'><div class='stat-num green'>{n_all_ok}</div>"
                f"<div class='stat-lbl'>All fields OK</div></div>",
                unsafe_allow_html=True,
            )
        with c4:
            col_c = "red" if n_issues else "green"
            st.markdown(
                f"<div class='stat-box'><div class='stat-num {col_c}'>{n_issues}</div>"
                f"<div class='stat-lbl'>Need review</div></div>",
                unsafe_allow_html=True,
            )

        if n_issues == 0:
            st.success("✅ All invoices passed validation.")
        else:
            st.warning(f"⚠️ {n_issues} invoice(s) need review — see the Excel report below.")

        # ── Export to Excel ────────────────────────────────────────────────────
        st.divider()
        export_rows = []
        for r in val_results:
            export_rows.append({
                "File":                  r["filename"],
                "Invoice No":            r["invoice_no"],
                "Found in Audit":        "Yes" if r["found"] else "No",
                "Voucher No":            r.get("voucher_aud") or "",
                "Inv Date (PDF)":        _fmt_date(r["inv_date"]),
                "Inv Date (Audit)":      _fmt_date(r["inv_date_aud"]),
                "Inv Date OK":           _icon(r["inv_date_ok"]),
                "Payment Date (Calc)":   _fmt_date(r["exp_due"]),
                "Payment Date (Audit)":  _fmt_date(r["due_date_aud"]),
                "Payment Date OK":       _icon(r["due_date_ok"]),
                "CC (PDF)":              r["cc_ext"]    or "",
                "CC (Audit)":            r["cc_aud"]    or "",
                "CC OK":                 _icon(r["cc_ok"]),
                "Vendor (PDF)":          r["vendor_ext"] or "",
                "Vendor (Audit)":        r["vendor_aud"] or "",
                "Vendor OK":             _icon(r["vendor_ok"]),
                "GL (PDF)":              r["gl_ext"]    or "",
                "GL (Audit)":            r["gl_aud"]    or "",
                "GL OK":                 _icon(r["gl_ok"]),
                "Net/Subtotal (PDF)":    _fmt_amt(r["inv_net"]),
                "Taxes GST+QST (PDF)":  _fmt_amt(r["inv_taxes"]),
                "Total (PDF)":          _fmt_amt(r["inv_total"]),
                "Total (Audit)":        _fmt_amt(r["aud_total"]),
                "Total OK":             _icon(r["total_ok"]),
            })
        df_exp = pd.DataFrame(export_rows)
        buf_xl = BytesIO()
        with pd.ExcelWriter(buf_xl, engine="openpyxl") as xl_writer:
            df_exp.to_excel(xl_writer, index=False)
        buf_xl.seek(0)
        st.download_button(
            "⬇️ Download Validation Report (Excel)",
            data=buf_xl.read(),
            file_name=f"ap_audit_validation_{date.today().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


# ══════════════════════════════════════════════════════════════════════════════
# TAB — STATEMENT RECONCILIATION
# ══════════════════════════════════════════════════════════════════════════════
if active_module == "recon":
    st.subheader("🧮 Statement Reconciliation")
    st.markdown(
        "Reconcile a vendor's **statement of account** against the accounting system "
        "and the invoice-copy tracking Excel. The tool walks every invoice on the "
        "statement and checks: is it registered in the system, does the amount match, "
        "do we have a copy on file, and does the PO on the statement match the PO on "
        "the file name."
    )

    vendor_keys = list(RECON_VENDORS.keys())
    recon_vendor = st.selectbox(
        "1️⃣ Vendor to reconcile",
        vendor_keys,
        format_func=lambda k: RECON_VENDORS[k]["label"],
    )
    st.caption(
        "The system extract and the invoice-copy tracking Excel are already scoped to "
        "this vendor when you export them — no extra filtering by name is applied; the "
        "vendor selection only decides which template is used to read the **statement**."
    )

    st.markdown("**2️⃣ Attach supporting files**")
    col_a, col_b, col_c = st.columns(3)
    with col_a:
        recon_statement_files = st.file_uploader(
            "📄 Statement of Account (PDF or Excel)",
            type=["pdf", "xlsx"],
            accept_multiple_files=True,
            key=f"recon_statement_{st.session_state.recon_upload_key}",
        )
        st.caption(
            "Upload more than one file if the vendor splits its statement "
            "(e.g. one per branch) — they'll be combined and reconciled "
            "together against a single system extract."
        )
    with col_b:
        recon_system_file = st.file_uploader(
            "🗄️ System extract (Excel)",
            type=["xlsx"],
            key=f"recon_system_{st.session_state.recon_upload_key}",
        )
    with col_c:
        recon_extraction_file = st.file_uploader(
            "📁 Invoice-copy tracking Excel (optional)",
            type=["xlsx"],
            key=f"recon_extraction_{st.session_state.recon_upload_key}",
        )

    col_run, col_clear = st.columns([1, 5])
    with col_run:
        do_recon = st.button(
            "🧮 Reconcile", type="primary",
            disabled=not (recon_statement_files and recon_system_file),
        )
    with col_clear:
        if st.button("🗑️ Clear", key="recon_clear"):
            st.session_state.recon_results = None
            st.session_state.recon_zip = None
            st.session_state.recon_upload_key += 1
            st.rerun()

    if do_recon:
        try:
            vendor_cfg = RECON_VENDORS[recon_vendor]
            check_po = vendor_cfg.get("check_po", True)
            statement = []
            grand_total = 0.0
            grand_total_complete = True
            for f in recon_statement_files:
                file_records, file_grand_total = vendor_cfg["parse_statement"](f.name, f.read())
                statement.extend(file_records)
                if file_grand_total is None:
                    grand_total_complete = False
                else:
                    grand_total += file_grand_total
            grand_total = grand_total if grand_total_complete else None
            system = parse_system_extract(recon_system_file.read())
            extraction = (
                parse_extraction_excel(recon_extraction_file.read())
                if recon_extraction_file else {}
            )
            if not statement:
                st.error("Could not find any invoice lines in the statement(s). Check the file format.")
            else:
                buckets = reconcile_statement(statement, system, extraction, check_po)
                statement_total = sum((r.get("amount") or 0) for r in statement)
                st.session_state.recon_results = buckets
                st.session_state.recon_statement_total = statement_total
                st.session_state.recon_grand_total = grand_total
                st.session_state.recon_check_po = check_po
                st.session_state.recon_zip = make_recon_report(buckets, statement_total, grand_total, check_po)
        except Exception as e:
            st.error(f"Error during reconciliation: {e}")

    buckets = st.session_state.recon_results
    if buckets:
        statement_total = st.session_state.recon_statement_total
        grand_total = st.session_state.recon_grand_total
        check_po = st.session_state.recon_check_po
        if grand_total is not None:
            diff = round(statement_total - grand_total, 2)
            if abs(diff) <= 0.02:
                st.success(
                    f"✅ Statement total matches: file says **CAD {grand_total:,.2f}**, "
                    f"parsed lines add up to **CAD {statement_total:,.2f}**."
                )
            else:
                st.error(
                    f"⚠️ Statement total does **not** match: file says **CAD {grand_total:,.2f}**, "
                    f"parsed lines add up to **CAD {statement_total:,.2f}** (difference "
                    f"**CAD {diff:,.2f}**) — the file may not have been read correctly."
                )
        else:
            st.warning(
                f"Could not find a total figure on the statement to check against — "
                f"parsed lines add up to **CAD {statement_total:,.2f}**."
            )

        st.markdown("<br>", unsafe_allow_html=True)
        summary_df = pd.DataFrame(
            [
                (label,
                 "" if count is None else f"{count:,}",
                 "" if amount is None else f"{amount:,.2f}")
                for label, count, amount in build_recon_summary_rows(buckets, statement_total, check_po)
            ],
            columns=["Category", "Invoice Count", "Amount (CAD)"],
        )
        st.dataframe(summary_df, use_container_width=True, hide_index=True)

        st.download_button(
            "⬇️ Download Reconciliation Report (Excel)",
            data=st.session_state.recon_zip,
            file_name=f"{recon_vendor}_reconciliation_{date.today().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=False,
            type="primary",
        )
        st.caption(
            "Full invoice-level detail (including the **has copy** and **PO mismatch** "
            "columns on the Pending Registration sheet) is in the Excel report."
        )
    elif not (recon_statement_files or recon_system_file):
        st.info("📂 Upload the statement and the system extract to run the reconciliation.")


# ══════════════════════════════════════════════════════════════════════════════
# TAB 6 — PAYMENT PACKAGER
# ══════════════════════════════════════════════════════════════════════════════
if active_module == "payment":
    st.subheader("💳 Payment Packager — Build the invoice packages for a payment run")
    st.markdown(
        "This tool works entirely through **upload / download** — it does **not** need access "
        "to your computer's folders or network drives, so it works from any browser without "
        "installing anything. Invoices are matched by **invoice number**: the first **8 "
        "characters** of each PDF's filename, against the **Reference** column (column **E**) "
        "of the AP payment report.\n\n"
        "1. Upload the **payment report** (Excel/CSV) — the list of invoices to pay.\n"
        "2. Add each **unpaid folder** separately (its name/vendor + all its current PDFs, "
        "paid or not) — one folder at a time, as many as you need.\n"
        "3. Click **Build payment packages** — you'll get a ZIP for the **Paid folder (AP / "
        "Vendors)**, a ZIP for the **payment-number folder (Finance)**, and — for each unpaid "
        "folder you added — a ready-to-swap-in **updated unpaid ZIP** with the paid invoices "
        "already removed, so you replace the whole folder instead of hunting files to delete."
    )

    def _pay_clean_ref(v) -> str:
        if v is None:
            return ""
        if isinstance(v, float) and v.is_integer():
            v = int(v)
        s = str(v).strip()
        if not s or s.lower() == "nan":
            return ""
        if re.fullmatch(r"\d+\.0+", s):
            s = s.split(".")[0]
        return s

    # ── 1. Payment report ────────────────────────────────────────────────────
    st.markdown("**1. Payment report (invoices to pay)**")
    pay_list_file = st.file_uploader(
        "Excel or CSV with the invoices selected for payment",
        type=["xlsx", "xls", "csv"],
        key="pay_list_upload",
    )

    pay_invoices = []
    if pay_list_file:
        df_pay = None
        try:
            if pay_list_file.name.lower().endswith(".csv"):
                df_pay = pd.read_csv(pay_list_file)
            else:
                df_pay = pd.read_excel(pay_list_file)
        except Exception as e:
            st.error(f"Could not read the file: {e}")

        if df_pay is not None and len(df_pay.columns):
            guess_idx = 4 if len(df_pay.columns) > 4 else 0
            for i, c in enumerate(df_pay.columns):
                if re.search(r"reference|referencia", str(c), re.I):
                    guess_idx = i
                    break
            pay_col = st.selectbox(
                "Column with the invoice number (defaults to column E — Reference)",
                options=list(df_pay.columns),
                index=guess_idx,
                key="pay_col_select",
            )
            st.caption(
                "Compared against the first 8 characters of each PDF's filename — "
                "e.g. `92006209_ML_24289954.pdf` → `92006209`."
            )
            pay_invoices = [
                _pay_clean_ref(v) for v in df_pay[pay_col].tolist()
            ]
            pay_invoices = [v for v in pay_invoices if v]
            st.caption(f"{len(pay_invoices)} invoice(s) loaded from **{pay_col}**")

    st.divider()

    # ── 2. Unpaid folders, added one at a time ──────────────────────────────
    st.markdown("**2. Unpaid folders**")
    st.caption(
        "Add every unpaid folder that might contain a selected invoice. Upload **all** the PDFs "
        "currently in that folder (not just the ones being paid) so the tool can hand back a "
        "complete, ready-to-swap-in replacement for it. If you upload everything in one single "
        "batch instead of folder by folder, any label works (e.g. `Unpaid`) — it's only used to "
        "name the ZIP you get back, it doesn't affect the matching."
    )

    existing_labels = {b["label"].lower() for b in st.session_state.payment_batches}
    _pay_form_key = st.session_state.payment_batch_form_key

    if st.session_state.payment_batches:
        _pay_total_mb = sum(
            len(f["bytes"]) for b in st.session_state.payment_batches for f in b["files"]
        ) / (1024 * 1024)
        cap_col, clr_col = st.columns([5, 1])
        with cap_col:
            st.caption(f"📦 {_pay_total_mb:,.1f} MB currently held in this session across all folders.")
        with clr_col:
            if st.button("🗑️ Clear all", key="pay_clear_all", use_container_width=True):
                st.session_state.payment_batches = []
                st.session_state.payment_result = None
                st.rerun()

    fc1, fc2 = st.columns([1, 2])
    with fc1:
        batch_label = st.text_input(
            "Folder name / vendor No.",
            placeholder="0101000430",
            key=f"pay_batch_label_{_pay_form_key}",
        )
    with fc2:
        batch_files = st.file_uploader(
            "All PDFs currently in that folder", type=["pdf"], accept_multiple_files=True,
            key=f"pay_batch_files_{_pay_form_key}",
        )
    add_batch = st.button("➕ Add this folder")

    if add_batch:
        label = batch_label.strip()
        if not label:
            st.warning("Enter a folder name / vendor number first.")
        elif label.lower() in existing_labels:
            st.warning(f"'{label}' was already added — remove it below first if you want to replace it.")
        elif not batch_files:
            st.warning("Upload that folder's PDFs first.")
        else:
            st.session_state.payment_batches.append({
                "label": label,
                "files": [{"name": f.name, "bytes": f.getvalue()} for f in batch_files],
            })
            st.session_state.payment_batch_form_key += 1
            st.rerun()

    if st.session_state.payment_batches:
        for i, b in enumerate(st.session_state.payment_batches):
            bc1, bc2 = st.columns([6, 1])
            with bc1:
                st.write(f"📁 **{b['label']}** — {len(b['files'])} file(s)")
            with bc2:
                if st.button("🗑️", key=f"pay_batch_del_{i}"):
                    st.session_state.payment_batches.pop(i)
                    st.rerun()
    else:
        st.info("No unpaid folders added yet.")

    st.divider()

    # ── 3. Payment number ─────────────────────────────────────────────────────
    st.markdown("**3. Payment number**")
    payment_no_txt = st.text_input(
        "Used to name the Finance package/folder",
        value="",
        placeholder="e.g. PAY-2026-08-17-001",
        key="pay_number_input",
    )

    st.divider()

    can_build = bool(pay_invoices and st.session_state.payment_batches and payment_no_txt.strip())
    if st.button("📦 Build payment packages", type="primary", disabled=not can_build):
        # Index every uploaded file across all folders, by invoice No. (first 8 chars of filename).
        upload_index = {}
        for b in st.session_state.payment_batches:
            for f in b["files"]:
                inv_key = f["name"][:8]
                upload_index.setdefault(inv_key, []).append(f)

        rows = []
        for inv_no in pay_invoices:
            matches = upload_index.get(inv_no, [])
            if len(matches) == 1:
                status = "found"
            elif len(matches) > 1:
                status = "duplicate"
            else:
                status = "not_found"
            rows.append({"invoice_no": inv_no, "matches": matches, "status": status})

        to_pack = [r for r in rows if r["status"] == "found"]
        paid_keys = {r["invoice_no"] for r in to_pack}

        # Paid and Finance folders get the exact same set of files — build the
        # ZIP once and reuse its bytes for both instead of duplicating it in memory.
        zip_paid_buf = BytesIO()
        with zipfile.ZipFile(zip_paid_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for r in to_pack:
                zf.writestr(r["matches"][0]["name"], r["matches"][0]["bytes"])
        zip_paid_bytes = zip_paid_buf.getvalue()

        payment_no_clean = payment_no_txt.strip()

        # Everything that will also go into the combined "download all" ZIP.
        all_entries = []
        for r in to_pack:
            all_entries.append((f"Facturas Pagadas (AP)/{r['matches'][0]['name']}", r["matches"][0]["bytes"]))
            all_entries.append((f"Finanzas - {payment_no_clean}/{r['matches'][0]['name']}", r["matches"][0]["bytes"]))

        # One "updated unpaid" ZIP per folder — everything except what just got paid.
        remaining_zips = []
        for b in st.session_state.payment_batches:
            remaining = [f for f in b["files"] if f["name"][:8] not in paid_keys]
            buf = BytesIO()
            with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
                for f in remaining:
                    zf.writestr(f["name"], f["bytes"])
            buf.seek(0)
            remaining_zips.append({
                "label": b["label"],
                "zip": buf.getvalue(),
                "removed": len(b["files"]) - len(remaining),
                "remaining": len(remaining),
            })
            safe_label = b["label"].replace("/", "-").replace("\\", "-")
            for f in remaining:
                all_entries.append((f"Unpaid - {safe_label}/{f['name']}", f["bytes"]))

        report_rows = [{
            "Invoice No":    r["invoice_no"],
            "File":          r["matches"][0]["name"] if len(r["matches"]) == 1 else
                              "; ".join(m["name"] for m in r["matches"]) if r["matches"] else "",
            "Status": {"found": "Packaged", "duplicate": "Duplicate — review",
                       "not_found": "Not found"}[r["status"]],
        } for r in rows]
        buf_report = BytesIO()
        with pd.ExcelWriter(buf_report, engine="openpyxl") as xl_writer:
            pd.DataFrame(report_rows).to_excel(xl_writer, index=False)
        buf_report.seek(0)
        all_entries.append((f"Checklist_{payment_no_clean}.xlsx", buf_report.getvalue()))

        zip_all_buf = BytesIO()
        with zipfile.ZipFile(zip_all_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for arcname, data in all_entries:
                zf.writestr(arcname, data)
        zip_all_buf.seek(0)

        st.session_state.payment_result = {
            "payment_no":     payment_no_clean,
            "rows":           [{
                "invoice_no": r["invoice_no"],
                "status":     r["status"],
                "files":      [m["name"] for m in r["matches"]],
            } for r in rows],
            "zip_paid":       zip_paid_bytes,
            "zip_finance":    zip_paid_bytes,
            "report_xlsx":    buf_report.getvalue(),
            "remaining_zips": remaining_zips,
            "zip_all":        zip_all_buf.getvalue(),
        }

    # ── Results ───────────────────────────────────────────────────────────────
    result = st.session_state.get("payment_result")
    if result:
        rows        = result["rows"]
        n_found     = sum(1 for r in rows if r["status"] == "found")
        n_dup       = sum(1 for r in rows if r["status"] == "duplicate")
        n_not_found = sum(1 for r in rows if r["status"] == "not_found")

        c1, c2, c3, c4 = st.columns(4)
        with c1:
            st.markdown(
                f"<div class='stat-box'><div class='stat-num blue'>{len(rows)}</div>"
                f"<div class='stat-lbl'>Selected</div></div>", unsafe_allow_html=True)
        with c2:
            st.markdown(
                f"<div class='stat-box'><div class='stat-num green'>{n_found}</div>"
                f"<div class='stat-lbl'>Packaged</div></div>", unsafe_allow_html=True)
        with c3:
            col_c = "amber" if n_dup else "green"
            st.markdown(
                f"<div class='stat-box'><div class='stat-num {col_c}'>{n_dup}</div>"
                f"<div class='stat-lbl'>Duplicate — review</div></div>", unsafe_allow_html=True)
        with c4:
            col_c = "red" if n_not_found else "green"
            st.markdown(
                f"<div class='stat-box'><div class='stat-num {col_c}'>{n_not_found}</div>"
                f"<div class='stat-lbl'>Not found</div></div>", unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        if n_dup:
            st.warning(
                "⚠️ Some invoice numbers matched more than one uploaded PDF (same first 8 "
                "characters, possibly in different folders) — those are skipped automatically "
                "(in both the paid package and every 'updated unpaid' ZIP) so the wrong file "
                "isn't moved. Resolve them manually, then rebuild."
            )
        if n_not_found:
            st.info(
                "ℹ️ Invoice numbers marked **Not found** weren't among the uploaded PDFs — check "
                "the number or upload the right folder and build the packages again."
            )

        if n_found:
            st.success(f"✅ {n_found} invoice(s) packaged.")

            st.download_button(
                "⬇️ Download everything (ZIP)",
                data=result["zip_all"],
                file_name=f"pago_{result['payment_no']}_completo.zip",
                mime="application/zip",
                type="primary",
                use_container_width=True,
            )
            st.caption(
                "Contains the Paid folder, the Finance folder, the checklist, and every "
                "updated unpaid folder — each in its own subfolder, ready to copy into place."
            )
            st.markdown("<br>", unsafe_allow_html=True)

            with st.expander("Or download each piece separately"):
                dcol1, dcol2, dcol3 = st.columns(3)
                with dcol1:
                    st.download_button(
                        "⬇️ Paid folder ZIP (AP / Vendors)",
                        data=result["zip_paid"],
                        file_name=f"facturas_pagadas_{date.today().strftime('%Y%m%d')}.zip",
                        mime="application/zip",
                        use_container_width=True,
                    )
                with dcol2:
                    st.download_button(
                        "⬇️ Finance ZIP (payment " + result["payment_no"] + ")",
                        data=result["zip_finance"],
                        file_name=f"pago_{result['payment_no']}.zip",
                        mime="application/zip",
                        use_container_width=True,
                    )
                with dcol3:
                    st.download_button(
                        "⬇️ Checklist (Excel)",
                        data=result["report_xlsx"],
                        file_name=f"payment_checklist_{result['payment_no']}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )

                st.markdown("**📁 Updated unpaid folders — replace each folder's contents with its ZIP**")
                for rz in result["remaining_zips"]:
                    safe_label = re.sub(r"[^A-Za-z0-9_.-]+", "_", rz["label"])
                    st.download_button(
                        f"⬇️ Unpaid — {rz['label']}  ({rz['removed']} removed, {rz['remaining']} remaining)",
                        data=rz["zip"],
                        file_name=f"unpaid_{safe_label}_{date.today().strftime('%Y%m%d')}.zip",
                        mime="application/zip",
                        key=f"pay_remaining_dl_{safe_label}",
                        use_container_width=True,
                    )

            st.info(
                "💡 **Next steps:** unzip the Paid ZIP into your **AP / Vendors paid** folder, "
                "unzip the Finance ZIP into the **payment-number** folder, then for each unpaid "
                "folder listed above, **delete everything inside it and unzip its updated ZIP in "
                "its place** — no need to search for and delete individual files."
            )


# ══════════════════════════════════════════════════════════════════════════════
# TAB 7 — DATABASE
# ══════════════════════════════════════════════════════════════════════════════
if active_module == "database":
    db_t1, db_t2 = st.tabs(["🏢 Vendors / Cost Centres", "📊 GL Accounts"])

    with db_t1:
        st.subheader("Vendor Table")
        st.caption("Maps the **prefix** (first 2 letters of the Order No.) → Vendor and Cost Centre")
        df_prov = pd.DataFrame(st.session_state.proveedores)
        edited_prov = st.data_editor(
            df_prov,
            column_config={
                "prefijo": st.column_config.TextColumn("Prefix", width="small",
                    help="First 2 letters of the customer Order No."),
                "vendor":  st.column_config.TextColumn("Vendor No.", width="medium"),
                "cc":      st.column_config.TextColumn("Cost Centre", width="medium"),
            },
            num_rows="dynamic",
            use_container_width=True,
            key="prov_editor",
        )
        if st.button("💾 Save changes — Vendors", type="primary"):
            new_prov = []
            for r in edited_prov.to_dict("records"):
                pref = r.get("prefijo")
                if pref and str(pref).strip() and str(pref).strip().upper() != "NAN":
                    new_prov.append({
                        "prefijo": str(pref).strip().upper(),
                        "vendor":  str(r.get("vendor", "")).strip(),
                        "cc":      str(r.get("cc", "")).strip(),
                    })
            st.session_state.proveedores = new_prov
            st.success(f"✅ Vendor table updated — {len(new_prov)} vendors saved")
            st.rerun()

    with db_t2:
        st.subheader("GL Codes Table")
        st.caption("Maps the **product code** from the invoice → GL Account")
        df_gl = pd.DataFrame(st.session_state.gl_codes)
        edited_gl = st.data_editor(
            df_gl,
            column_config={
                "codigo": st.column_config.TextColumn("Product Code", width="medium"),
                "gl":     st.column_config.TextColumn("GL Account", width="medium"),
            },
            num_rows="dynamic",
            use_container_width=True,
            key="gl_editor",
        )
        if st.button("💾 Save changes — GL", type="primary"):
            new_gl = []
            for r in edited_gl.to_dict("records"):
                cod = r.get("codigo")
                gl  = r.get("gl")
                if cod and str(cod).strip() and str(cod).strip().upper() != "NAN":
                    new_gl.append({
                        "codigo": str(cod).strip().upper(),
                        "gl":     str(gl).strip() if gl and str(gl).strip().upper() != "NAN" else "",
                    })
            st.session_state.gl_codes = new_gl
            st.success(f"✅ GL table updated — {len(new_gl)} codes saved")
            st.rerun()

        st.divider()
        st.subheader("📥 Import from Excel (maestro_contable.xlsx)")
        xl_up = st.file_uploader("Upload Excel file", type=["xlsx"], key="xl_import")
        if xl_up:
            try:
                wb = openpyxl.load_workbook(xl_up)
                imported_vendors = 0
                imported_gl = 0

                if "proveedores" in wb.sheetnames:
                    ws = wb["proveedores"]
                    new_proveedores = []
                    for r in ws.iter_rows(min_row=2, values_only=True):
                        if len(r) >= 3 and r[0] not in (None, ""):
                            new_proveedores.append({
                                "prefijo": str(r[0]).strip(),
                                "vendor":  str(r[1]).strip() if r[1] is not None else "",
                                "cc":      str(r[2]).strip() if r[2] is not None else "",
                            })
                    if new_proveedores:
                        st.session_state.proveedores = new_proveedores
                        imported_vendors = len(new_proveedores)

                if "cuentas_gl" in wb.sheetnames:
                    ws = wb["cuentas_gl"]
                    new_gl = []
                    for r in ws.iter_rows(min_row=2, values_only=True):
                        if len(r) >= 2 and r[0] not in (None, "") and r[1] not in (None, ""):
                            new_gl.append({
                                "codigo": str(r[0]).strip(),
                                "gl":     str(r[1]).strip(),
                            })
                    if new_gl:
                        st.session_state.gl_codes = new_gl
                        imported_gl = len(new_gl)

                if imported_vendors == 0 and imported_gl == 0:
                    st.warning(
                        "⚠️ No data imported. Make sure the Excel has sheets named "
                        "**'proveedores'** and/or **'cuentas_gl'** with data starting on row 2."
                    )
                else:
                    st.success(
                        f"✅ Imported: {imported_vendors} vendors, {imported_gl} GL codes"
                    )
                    st.rerun()
            except Exception as e:
                st.error(f"Error importing Excel: {e}")


# ══════════════════════════════════════════════════════════════════════════════
# TAB 8 — SETTINGS
# ══════════════════════════════════════════════════════════════════════════════
if active_module == "settings":
    st.subheader("General Settings")
    col1, col2 = st.columns(2)

    with col1:
        st.markdown("**👥 System Users**")
        users_txt = st.text_area(
            "One user per line:",
            value="\n".join(st.session_state.usuarios),
            height=130,
        )
        if st.button("💾 Save users"):
            new_users = [u.strip() for u in users_txt.splitlines() if u.strip()]
            if new_users:
                st.session_state.usuarios = new_users
                st.success(f"✅ {len(new_users)} user(s) saved")
            else:
                st.warning("The list cannot be empty")

    with col2:
        st.markdown("**📍 Stamp Position on PDF**")
        st.caption("Invoice PDFs are landscape (792 × 612 pts). The stamp has a white background with a red border.")

        c_x, c_y = st.columns(2)
        with c_x:
            new_sx = st.number_input("X — from left", 0, 780, st.session_state.stamp_x, step=5)
        with c_y:
            new_sy = st.number_input("Y — stamp top (from bottom)", 0, 610, st.session_state.stamp_y_top, step=5)

        c_w, c_h = st.columns(2)
        with c_w:
            new_sw = st.number_input("Width", 80, 400, st.session_state.stamp_w, step=5)
        with c_h:
            new_sh = st.number_input("Height", 40, 200, st.session_state.stamp_h, step=5)

        if st.button("💾 Save position"):
            st.session_state.stamp_x     = new_sx
            st.session_state.stamp_y_top = new_sy
            st.session_state.stamp_w     = new_sw
            st.session_state.stamp_h     = new_sh
            st.success("✅ Position updated")

        st.info(
            "💡 **Default values (centred top):** X=281, Y=594, Width=230, Height=82  \n"
            "Adjust if the stamp does not land in the correct area of the invoice."
        )

    st.divider()

    with st.expander("ℹ️ System Information"):
        st.markdown(f"""
        **Atlantic Invoice Tools v3.0**

        **Workflow:**
        1. **✂️ Split** — Upload Atlantic batch PDFs → auto-detect invoices → split into individual files
           named `{{Invoice No}} {{CC}} {{BOL}}.pdf`
        2. **🔗 Match** — Upload split invoices + PO PDFs → merge matched pairs
        3. **📤 Codify** — Stamp invoices with Vendor / CC / GL / Date

        **Split naming logic:**
        - Invoice No: from `INVOICE No/No DE FACTURE` field
        - CC: first 2 chars of Customer Order No. → lookup in Vendor table → first 2 chars of `cc` field
          *(e.g. MD → EV01 → **EV**)*
        - BOL: second line of Bill of Lading column if two lines exist, otherwise the single line value

        **Coding logic:**
        - The first **2 characters** of the Customer Order No. determine the CC prefix
        - The prefix is looked up in the **Vendor table** → Vendor + Cost Centre
        - The **Product Code** is extracted from the invoice and looked up in the **GL table** → GL Account
        - **Type-6 invoice exception:** if the invoice No. starts with `6` →
          Vendor = `{VENDOR_EXCEPCION}` (fixed), CC = manual user selection

        **Vendors in database:** {len(st.session_state.proveedores)}
        **GL codes in database:** {len(st.session_state.gl_codes)}
        **Users:** {', '.join(st.session_state.usuarios)}
        """)


# ══════════════════════════════════════════════════════════════════════════════
# TAB 9 — CONTROL FACTURAS (OTROS PROVEEDORES)
# ══════════════════════════════════════════════════════════════════════════════
if active_module == "control_prov":
    st.caption(
        "Recepción, codificación y control de proveedores **distintos a Atlantic** — "
        "carga manual de facturas (cada proveedor tiene su propio formato, así que no "
        "se intenta leer los PDF automáticamente), catálogo de proveedores y "
        "responsables, sello de codificación con posición ajustable a mano, y "
        "seguimiento de estado."
    )
    cf_tabs = st.tabs([
        "🧾 Recepción", "🏢 Proveedores", "👥 Responsables", "🏷️ Sello / Timbre",
        "📋 Control", "📌 Reglas especiales (GL/CC)", "🔄 Cruce con compras (AP)",
        "📥 Importar tu Excel actual",
    ])

    cf_columns = [
        "id", "fournisseur", "nom_system", "nro_vendor", "nro_facture", "date_facture",
        "division", "po_reception", "responsable", "etat", "problema", "comentario",
        "ultima_actualizacion", "date_reception", "poste", "payee", "nom_pdf", "cc", "gl", "monto",
    ]

    # ── Recepción ────────────────────────────────────────────────────────────
    with cf_tabs[0]:
        st.subheader("Recepción y codificación de facturas")
        st.caption("Agregá cada factura a mano — como en la hoja CONTROL del Excel.")

        prov_codes = sorted(
            (set(p.get("codigo", "") for p in st.session_state.cf_proveedores)
             | set(f.get("fournisseur", "") for f in st.session_state.cf_facturas))
            - {""}
        )
        resp_options = sorted(
            (set(st.session_state.cf_responsables)
             | set(f.get("responsable", "") for f in st.session_state.cf_facturas))
            - {""}
        )
        estado_options = sorted(
            (set(st.session_state.cf_estados)
             | set(f.get("etat", "") for f in st.session_state.cf_facturas))
            - {""}
        )

        with st.expander("➕ Agregar factura nueva", expanded=not st.session_state.cf_facturas):
            st.caption("Igual que el recuadro de carga del Excel — completá y tocá «Agregar».")
            with st.form("cf_quick_add_form", clear_on_submit=True):
                qa1, qa2, qa3 = st.columns(3)
                with qa1:
                    qa_prov = st.selectbox("Proveedor", [""] + prov_codes, key="cf_qa_prov")
                    qa_nro_facture = st.text_input("Nro. Factura", key="cf_qa_nro_facture")
                    qa_division = st.text_input("División", key="cf_qa_division")
                with qa2:
                    qa_date_facture = st.date_input("Fecha factura", value=date.today(), key="cf_qa_date_facture")
                    qa_po = st.text_input("PO / Recepción", key="cf_qa_po")
                    qa_monto = st.number_input("Monto + impuestos", min_value=0.0, step=0.01, key="cf_qa_monto")
                with qa3:
                    qa_resp = st.selectbox("Responsable", [""] + resp_options, key="cf_qa_resp")
                    qa_estado = st.selectbox("Estado", [""] + estado_options, key="cf_qa_estado")

                qa_p = _cf_proveedor_lookup(qa_prov) if qa_prov else None
                if qa_p:
                    st.caption(f"↳ {qa_p.get('nombre', '')} · # Vendor {qa_p.get('vendor_no', '')}")
                elif qa_prov:
                    st.caption("↳ Proveedor no encontrado en la base — completalo en la pestaña Proveedores.")
                st.caption("El **Nombre PDF** se arma solo: Proveedor + Nro. Factura + División + PO.")

                if st.form_submit_button("➕ Agregar (Ajouter)", type="primary"):
                    if not qa_prov or not qa_nro_facture.strip():
                        st.error("Completá al menos Proveedor y Nro. Factura.")
                    else:
                        today_str = date.today().isoformat()
                        qa_division_v = qa_division.strip()
                        qa_po_v = qa_po.strip()
                        new_row = {
                            "id":                   _cf_next_id(),
                            "fournisseur":          qa_prov,
                            "nom_system":           qa_p.get("nombre", "") if qa_p else "",
                            "nro_vendor":           qa_p.get("vendor_no", "") if qa_p else "",
                            "nro_facture":          qa_nro_facture.strip(),
                            "date_facture":         qa_date_facture.isoformat(),
                            "division":             qa_division_v,
                            "po_reception":         qa_po_v,
                            "responsable":          qa_resp,
                            "etat":                 qa_estado,
                            "problema":             False,
                            "comentario":           "",
                            "ultima_actualizacion": today_str,
                            "date_reception":       today_str,
                            "poste":                False,
                            "payee":                False,
                            "nom_pdf":              _cf_build_nom_pdf(qa_prov, qa_nro_facture.strip(), qa_division_v, qa_po_v),
                            "cc":                   "",
                            "gl":                   "",
                            "monto":                qa_monto or None,
                        }
                        st.session_state.cf_facturas.append(new_row)
                        st.success(f"✅ Factura #{new_row['id']} agregada — {qa_prov} · {qa_nro_facture.strip()}")
                        st.rerun()

        fcol1, fcol2, fcol3, fcol4 = st.columns(4)
        with fcol1:
            filt_prov = st.multiselect("Filtrar por proveedor", prov_codes, key="cf_filt_prov")
        with fcol2:
            filt_etat = st.multiselect("Filtrar por estado", estado_options, key="cf_filt_etat")
        with fcol3:
            filt_pend = st.checkbox("Solo pendientes de pago", key="cf_filt_pend")
        with fcol4:
            filt_problema = st.checkbox("Solo con problema", key="cf_filt_problema")

        rows = st.session_state.cf_facturas
        if filt_prov:
            rows = [r for r in rows if r.get("fournisseur") in filt_prov]
        if filt_etat:
            rows = [r for r in rows if r.get("etat") in filt_etat]
        if filt_pend:
            rows = [r for r in rows if not r.get("payee")]
        if filt_problema:
            rows = [r for r in rows if r.get("problema")]

        df = pd.DataFrame(rows, columns=cf_columns) if rows else pd.DataFrame(columns=cf_columns)

        edited = st.data_editor(
            df,
            column_config={
                "id":             st.column_config.NumberColumn("ID", disabled=True, width="small"),
                "fournisseur":    st.column_config.SelectboxColumn("Proveedor", options=prov_codes or [""]),
                "nom_system":     st.column_config.TextColumn("Nombre sistema"),
                "nro_vendor":     st.column_config.TextColumn("# Vendor"),
                "nro_facture":    st.column_config.TextColumn("Nro. Factura"),
                "date_facture":   st.column_config.TextColumn("Fecha factura (AAAA-MM-DD)"),
                "division":       st.column_config.TextColumn("División"),
                "po_reception":   st.column_config.TextColumn("PO / Recepción"),
                "responsable":    st.column_config.SelectboxColumn("Responsable", options=resp_options or [""]),
                "etat":           st.column_config.SelectboxColumn("Estado", options=estado_options or [""]),
                "problema":       st.column_config.CheckboxColumn("¿Problema?"),
                "comentario":     st.column_config.TextColumn("Comentario"),
                "ultima_actualizacion": st.column_config.TextColumn(
                    "Mis à jour", disabled=True, help="Se actualiza sola cuando cambia el Estado"
                ),
                "date_reception": st.column_config.TextColumn("Fecha recepción (AAAA-MM-DD)"),
                "poste":          st.column_config.CheckboxColumn("Posté"),
                "payee":          st.column_config.CheckboxColumn("Payée"),
                "nom_pdf":        st.column_config.TextColumn(
                    "Nombre PDF", disabled=True,
                    help="Se arma solo: Proveedor + Nro. Factura + División + PO",
                ),
                "cc":             st.column_config.TextColumn("CC"),
                "gl":             st.column_config.TextColumn("GL"),
                "monto":          st.column_config.NumberColumn("Monto + impuestos", format="%.2f"),
            },
            num_rows="dynamic",
            use_container_width=True,
            key="cf_facturas_editor",
            hide_index=True,
        )

        bcol1, bcol2 = st.columns([1, 2])
        with bcol1:
            if st.button("💾 Guardar cambios", type="primary", key="cf_save_facturas"):
                old_etat_by_id = {r["id"]: r.get("etat", "") for r in st.session_state.cf_facturas}
                today_str = date.today().isoformat()
                new_rows = []
                for r in edited.to_dict("records"):
                    if not str(r.get("fournisseur") or "").strip() and not str(r.get("nro_facture") or "").strip():
                        continue
                    rid = r.get("id")
                    rid = int(rid) if rid is not None and not pd.isna(rid) else _cf_next_id()
                    new_etat = r.get("etat", "")
                    # "Mis à jour" tracks changes in Estado — stamp today's
                    # date whenever it moved (including a brand new row's
                    # first status), otherwise keep whatever it already had.
                    ultima = today_str if (rid not in old_etat_by_id or old_etat_by_id[rid] != new_etat) \
                        else (r.get("ultima_actualizacion") or "")
                    new_rows.append({
                        **r, "id": rid,
                        "poste": bool(r.get("poste")),
                        "payee": bool(r.get("payee")),
                        "problema": bool(r.get("problema")),
                        "ultima_actualizacion": ultima,
                        # Nombre PDF is never typed by hand — always rebuilt
                        # from the row's own Proveedor/Factura/División/PO.
                        "nom_pdf": _cf_build_nom_pdf(
                            r.get("fournisseur"), r.get("nro_facture"), r.get("division"), r.get("po_reception")
                        ),
                    })
                st.session_state.cf_facturas = new_rows
                st.success(f"✅ {len(new_rows)} facturas guardadas")
                st.rerun()
        with bcol2:
            if st.button("🔎 Autocompletar nombre / CC / GL", key="cf_autofill"):
                n_name, n_gl = 0, 0
                for r in st.session_state.cf_facturas:
                    p = _cf_proveedor_lookup(r.get("fournisseur", ""))
                    if p and not r.get("nom_system"):
                        r["nom_system"] = p.get("nombre", "")
                        n_name += 1
                    if p and not r.get("nro_vendor"):
                        r["nro_vendor"] = p.get("vendor_no", "")
                    if not r.get("gl"):
                        for reg in st.session_state.cf_reglas:
                            if reg.get("proveedor") == r.get("fournisseur") and reg.get("cc") == (r.get("division") or r.get("cc")):
                                r["cc"] = r.get("cc") or reg.get("cc", "")
                                r["gl"] = reg.get("gl", "")
                                n_gl += 1
                                break
                st.success(f"✅ {n_name} nombres y {n_gl} GL/CC completados")
                st.rerun()

        st.divider()
        m1, m2, m3, m4, m5 = st.columns(5)
        m1.metric("Total facturas", len(st.session_state.cf_facturas))
        m2.metric("Pendientes de pago", sum(1 for r in st.session_state.cf_facturas if not r.get("payee")))
        m3.metric("Con problema", sum(1 for r in st.session_state.cf_facturas if r.get("problema")))
        m4.metric("Sin PDF registrado", sum(1 for r in st.session_state.cf_facturas if not r.get("nom_pdf")))
        total_monto = sum(r.get("monto") or 0 for r in st.session_state.cf_facturas)
        m5.metric("Monto total", f"${total_monto:,.2f}")

    # ── Proveedores ──────────────────────────────────────────────────────────
    with cf_tabs[1]:
        st.subheader("Base de proveedores")
        st.caption(
            "Reemplaza la hoja **FOURNISSEUR** del Excel — nombre corto, nombre de "
            "sistema y número de proveedor. Agregá una fila nueva con el **+** de abajo "
            "de la tabla."
        )
        prov_cols = ["codigo", "nombre", "vendor_no", "divisiones", "activo", "notas"]
        df_p = pd.DataFrame(st.session_state.cf_proveedores, columns=prov_cols) \
            if st.session_state.cf_proveedores else pd.DataFrame(columns=prov_cols)
        edited_p = st.data_editor(
            df_p,
            column_config={
                "codigo":     st.column_config.TextColumn("Nombre corto", help="Código corto usado en Recepción, ej. 'D.Proden'"),
                "nombre":     st.column_config.TextColumn("Nombre de sistema", help="Razón social tal como figura en el sistema contable"),
                "vendor_no":  st.column_config.TextColumn("# Proveedor", help="Número de proveedor — solo dígitos (se guarda como texto para no perder ceros a la izquierda)"),
                "divisiones": st.column_config.TextColumn("Divisiones (separadas por coma)"),
                "activo":     st.column_config.CheckboxColumn("Activo"),
                "notas":      st.column_config.TextColumn("Notas"),
            },
            num_rows="dynamic", use_container_width=True, key="cf_prov_editor", hide_index=True,
        )
        if st.button("💾 Guardar cambios — Proveedores", type="primary", key="cf_save_prov"):
            new_p, non_numeric = [], []
            for r in edited_p.to_dict("records"):
                cod = r.get("codigo")
                if cod and str(cod).strip():
                    vendor_no = str(r.get("vendor_no") or "").strip()
                    if vendor_no and not vendor_no.isdigit():
                        non_numeric.append(f"{cod} ({vendor_no})")
                    new_p.append({
                        "codigo":     str(cod).strip(),
                        "nombre":     str(r.get("nombre") or "").strip(),
                        "vendor_no":  vendor_no,
                        "divisiones": str(r.get("divisiones") or "").strip(),
                        "activo":     bool(r.get("activo", True)),
                        "notas":      str(r.get("notas") or "").strip(),
                    })
            st.session_state.cf_proveedores = new_p
            st.success(f"✅ {len(new_p)} proveedores guardados")
            if non_numeric:
                st.warning("⚠️ El # Proveedor debería ser solo números — revisá: " + ", ".join(non_numeric))
            st.rerun()

    # ── Responsables ─────────────────────────────────────────────────────────
    with cf_tabs[2]:
        st.subheader("Responsables")
        st.caption(
            "Personas que pueden quedar asignadas a una factura en Recepción y Control. "
            "Agregá o quitá nombres según haga falta."
        )
        df_resp = pd.DataFrame({"nombre": st.session_state.cf_responsables}) \
            if st.session_state.cf_responsables else pd.DataFrame(columns=["nombre"])
        edited_resp = st.data_editor(
            df_resp,
            column_config={"nombre": st.column_config.TextColumn("Nombre")},
            num_rows="dynamic", use_container_width=True, key="cf_resp_editor", hide_index=True,
        )
        if st.button("💾 Guardar cambios — Responsables", type="primary", key="cf_save_resp"):
            new_resp = sorted({
                str(r.get("nombre") or "").strip()
                for r in edited_resp.to_dict("records") if str(r.get("nombre") or "").strip()
            })
            st.session_state.cf_responsables = new_resp
            st.success(f"✅ {len(new_resp)} responsables guardados")
            st.rerun()

    # ── Sello / Timbre ───────────────────────────────────────────────────────
    with cf_tabs[3]:
        st.subheader("Crear y colocar el sello de codificación")
        st.caption(
            "Genera el mismo sello rojo de la hoja CONTROL (Vendor / CC-GL / Periodo / "
            "Prix A/T / PO-Recepción / Posted By) y lo coloca sobre la factura. Cada "
            "proveedor tiene un formato distinto, así que la posición se ajusta a mano "
            "con los controles — mirá la vista previa para encontrar un espacio en blanco."
        )

        st.markdown("**1. Datos del sello**")
        picked_factura = None
        if st.session_state.cf_facturas:
            use_existing = st.checkbox(
                "Completar con los datos de una factura ya registrada", value=True, key="cf_stamp_use_existing"
            )
        else:
            use_existing = False

        vendor_default = cc_default = gl_default = po_default = ""
        monto_default = ""
        resp_default = current_user
        sel_key = "manual"

        if use_existing:
            options = {
                f"#{f['id']} — {f.get('fournisseur', '')} · {f.get('nro_facture', '')} ({f.get('date_facture', '')})": f
                for f in sorted(st.session_state.cf_facturas, key=lambda f: f["id"])
            }
            sel_label = st.selectbox("Factura (por correlativo #)", list(options.keys()), key="cf_stamp_pick")
            picked_factura = options[sel_label]
            sel_key = picked_factura["id"]
            vendor_default = picked_factura.get("nro_vendor", "")
            cc_default      = picked_factura.get("cc", "")
            gl_default      = picked_factura.get("gl", "")
            po_default      = picked_factura.get("po_reception", "")
            monto_default   = picked_factura.get("monto") or ""
            resp_default    = picked_factura.get("responsable") or current_user

        # Keying each field by the selected factura's id (not a fixed key)
        # forces a fresh widget — with the new defaults — every time the
        # picked factura changes, instead of Streamlit keeping whatever the
        # field held for the previously selected one.
        s1, s2, s3 = st.columns(3)
        with s1:
            stamp_vendor = st.text_input("Vendor", value=str(vendor_default), key=f"cf_stamp_vendor_{sel_key}")
            stamp_cc = st.text_input("CC", value=str(cc_default), key=f"cf_stamp_cc_{sel_key}")
        with s2:
            stamp_gl = st.text_input("GL", value=str(gl_default), key=f"cf_stamp_gl_{sel_key}")
            stamp_po = st.text_input("PO / Recepción", value=str(po_default), key=f"cf_stamp_po_{sel_key}")
        with s3:
            stamp_periode = st.text_input("Periodo (MM - AAAA)", value=date.today().strftime("%m - %Y"), key="cf_stamp_periode")
            stamp_monto = st.text_input("Prix A/T", value=str(monto_default), key=f"cf_stamp_monto_{sel_key}")
        stamp_posted_by = st.text_input("Posted By", value=str(resp_default), key=f"cf_stamp_posted_by_{sel_key}")

        cc_gl = f"{stamp_cc} - {stamp_gl}" if stamp_gl else stamp_cc
        stamp_lines = [
            f"VENDOR: {stamp_vendor}",
            f"CC - GL: {cc_gl}",
            f"PERIODE: {stamp_periode}",
            f"PRIX A/T: {stamp_monto}",
            f"PO-RECEPTION: {stamp_po}",
            f"POSTED BY: {stamp_posted_by}",
        ]

        st.markdown("**2. Factura a sellar**")
        stamp_pdf_up = st.file_uploader("PDF de la factura", type=["pdf"], key="cf_stamp_pdf_uploader")

        if stamp_pdf_up:
            pdf_bytes = stamp_pdf_up.getvalue()
            page = PdfReader(BytesIO(pdf_bytes)).pages[0]
            page_w = float(page.mediabox.width)
            page_h = float(page.mediabox.height)

            st.markdown("**3. Ajustá la posición del sello (buscá un espacio en blanco)**")
            p1, p2 = st.columns(2)
            with p1:
                stamp_x = st.slider("X — desde la izquierda", 0, int(page_w),
                                     min(st.session_state.cf_stamp_x, int(page_w)), key="cf_stamp_x_sl")
                stamp_w = st.slider("Ancho", 120, 400, st.session_state.cf_stamp_w, key="cf_stamp_w_sl")
            with p2:
                stamp_y = st.slider("Y — desde arriba", 0, int(page_h),
                                     min(st.session_state.cf_stamp_y, int(page_h)), key="cf_stamp_y_sl")
                stamp_h = st.slider("Alto", 70, 220, st.session_state.cf_stamp_h, key="cf_stamp_h_sl")

            try:
                preview_img = _cf_render_stamp_preview(pdf_bytes, stamp_x, stamp_y, stamp_w, stamp_h)
                st.image(preview_img, caption="Vista previa del sello sobre la factura", use_container_width=True)
            except Exception as e:
                st.warning(f"No se pudo generar la vista previa ({e}) — igual podés generar la factura sellada.")

            update_registro = False
            if picked_factura is not None:
                update_registro = st.checkbox(
                    f"Al generar, marcar esta factura como «Sellada - enviada a aprobación»",
                    value=True, key="cf_stamp_update_estado",
                )

            if st.button("🏷️ Generar factura sellada", type="primary", key="cf_stamp_generate"):
                stamp_bytes = _cf_create_stamp_pdf(stamp_lines, stamp_x, stamp_y, stamp_w, stamp_h, page_w, page_h)
                stamped = stamp_pdf(pdf_bytes, stamp_bytes)
                st.session_state.cf_stamp_x, st.session_state.cf_stamp_y = stamp_x, stamp_y
                st.session_state.cf_stamp_w, st.session_state.cf_stamp_h = stamp_w, stamp_h
                st.session_state.cf_stamped_result = {
                    "bytes": stamped,
                    "name": f"SELLADA_{re.sub(r'[^A-Za-z0-9_.-]+', '_', stamp_pdf_up.name)}",
                }
                if update_registro and picked_factura is not None:
                    picked_factura["etat"] = "Sellada - enviada a aprobación"
                    picked_factura["ultima_actualizacion"] = date.today().isoformat()
                    if "Sellada - enviada a aprobación" not in st.session_state.cf_estados:
                        st.session_state.cf_estados.append("Sellada - enviada a aprobación")
                st.rerun()

        if st.session_state.cf_stamped_result:
            res = st.session_state.cf_stamped_result
            st.download_button(
                "⬇️ Descargar factura sellada", data=res["bytes"], file_name=res["name"],
                mime="application/pdf", key="cf_stamp_download", type="primary",
            )

    # ── Control ──────────────────────────────────────────────────────────────
    with cf_tabs[4]:
        st.subheader("Control y seguimiento de estado")
        st.caption(
            "Validá las facturas recibidas: filtrá y cambiá el estado o marcá si tienen "
            "algún problema. Reemplaza la hoja VALIDATION del Excel."
        )

        cprov = sorted(set(f.get("fournisseur", "") for f in st.session_state.cf_facturas) - {""})
        cestados = sorted((set(st.session_state.cf_estados)
                            | set(f.get("etat", "") for f in st.session_state.cf_facturas)) - {""})
        cresp = sorted((set(st.session_state.cf_responsables)
                         | set(f.get("responsable", "") for f in st.session_state.cf_facturas)) - {""})

        cc1, cc2, cc3 = st.columns(3)
        with cc1:
            c_filt_prov = st.multiselect("Proveedor", cprov, key="cf_ctrl_filt_prov")
        with cc2:
            c_filt_estado = st.multiselect("Estado", cestados, key="cf_ctrl_filt_estado")
        with cc3:
            c_filt_problema = st.checkbox("Solo con problema", key="cf_ctrl_filt_problema")

        rows = st.session_state.cf_facturas
        if c_filt_prov:
            rows = [r for r in rows if r.get("fournisseur") in c_filt_prov]
        if c_filt_estado:
            rows = [r for r in rows if r.get("etat") in c_filt_estado]
        if c_filt_problema:
            rows = [r for r in rows if r.get("problema")]

        ctrl_cols = ["id", "fournisseur", "nro_facture", "date_reception", "responsable",
                     "etat", "problema", "comentario", "ultima_actualizacion"]
        df_ctrl = pd.DataFrame(rows, columns=ctrl_cols) if rows else pd.DataFrame(columns=ctrl_cols)

        edited_ctrl = st.data_editor(
            df_ctrl,
            column_config={
                "id":             st.column_config.NumberColumn("ID", disabled=True, width="small"),
                "fournisseur":    st.column_config.TextColumn("Proveedor", disabled=True),
                "nro_facture":    st.column_config.TextColumn("Nro. Factura", disabled=True),
                "date_reception": st.column_config.TextColumn("Fecha recepción", disabled=True),
                "responsable":    st.column_config.SelectboxColumn("Responsable", options=cresp or [""]),
                "etat":           st.column_config.SelectboxColumn("Estado", options=cestados or [""]),
                "problema":       st.column_config.CheckboxColumn("¿Problema?"),
                "comentario":     st.column_config.TextColumn("Comentario"),
                "ultima_actualizacion": st.column_config.TextColumn(
                    "Mis à jour", disabled=True, help="Se actualiza sola cuando cambia el Estado"
                ),
            },
            use_container_width=True, hide_index=True, key="cf_control_editor",
        )

        if st.button("💾 Guardar cambios de estado", type="primary", key="cf_save_control"):
            by_id = {r["id"]: r for r in st.session_state.cf_facturas}
            today_str = date.today().isoformat()
            n = 0
            for r in edited_ctrl.to_dict("records"):
                rid = r.get("id")
                if rid in by_id:
                    if by_id[rid].get("etat", "") != r.get("etat", ""):
                        by_id[rid]["ultima_actualizacion"] = today_str
                    by_id[rid]["responsable"] = r.get("responsable", "")
                    by_id[rid]["etat"] = r.get("etat", "")
                    by_id[rid]["problema"] = bool(r.get("problema"))
                    by_id[rid]["comentario"] = r.get("comentario", "")
                    n += 1
            st.success(f"✅ {n} facturas actualizadas")
            st.rerun()

        st.divider()
        cm1, cm2, cm3 = st.columns(3)
        cm1.metric("Mostrando", len(rows))
        cm2.metric("Con problema", sum(1 for r in st.session_state.cf_facturas if r.get("problema")))
        cm3.metric("Sin estado asignado", sum(1 for r in st.session_state.cf_facturas if not r.get("etat")))

        with st.expander("➕ Agregar un nuevo estado a la lista"):
            new_estado = st.text_input("Nuevo estado", key="cf_new_estado_input")
            if st.button("Agregar", key="cf_add_estado_btn") and new_estado.strip():
                if new_estado.strip() not in st.session_state.cf_estados:
                    st.session_state.cf_estados.append(new_estado.strip())
                    st.success(f"✅ Estado «{new_estado.strip()}» agregado")
                    st.rerun()

    # ── Reglas especiales ────────────────────────────────────────────────────
    with cf_tabs[5]:
        st.subheader("Reglas especiales de GL / CC por proveedor")
        st.caption(
            "Generaliza la hoja **Xerox** del Excel a cualquier proveedor que tenga "
            "varios GL/CC según concepto o división."
        )
        reglas_cols = ["proveedor", "concepto", "gl", "cc", "no_cliente", "notas"]
        df_r = pd.DataFrame(st.session_state.cf_reglas, columns=reglas_cols) \
            if st.session_state.cf_reglas else pd.DataFrame(columns=reglas_cols)
        edited_r = st.data_editor(
            df_r,
            column_config={
                "proveedor":  st.column_config.TextColumn("Proveedor"),
                "concepto":   st.column_config.TextColumn("Concepto"),
                "gl":         st.column_config.TextColumn("GL"),
                "cc":         st.column_config.TextColumn("CC"),
                "no_cliente": st.column_config.TextColumn("No. Cliente"),
                "notas":      st.column_config.TextColumn("Notas"),
            },
            num_rows="dynamic", use_container_width=True, key="cf_reglas_editor", hide_index=True,
        )
        if st.button("💾 Guardar cambios — Reglas", type="primary", key="cf_save_reglas"):
            new_r = []
            for r in edited_r.to_dict("records"):
                if str(r.get("proveedor") or "").strip():
                    new_r.append({k: str(r.get(k) or "").strip() for k in reglas_cols})
            st.session_state.cf_reglas = new_r
            st.success(f"✅ {len(new_r)} reglas guardadas")
            st.rerun()

    # ── Cruce con compras (AP) ───────────────────────────────────────────────
    with cf_tabs[6]:
        st.subheader("Cruzar contra el archivo de compras (AP)")
        st.caption(
            "Reemplaza el link externo del Excel a `Fichier des achats - AP.xlsx` (105k filas). "
            "Subí ese archivo cada vez que lo actualices — se usa solo para completar el "
            "PO/Recepción de las facturas pendientes. **El archivo no se guarda**, solo el "
            "resultado del cruce queda en el registro."
        )
        ap_up = st.file_uploader("Excel de compras (AP)", type=["xlsx", "xls"], key="cf_ap_uploader")
        if ap_up:
            ap_bytes = ap_up.read()
            try:
                wb_ap = openpyxl.load_workbook(BytesIO(ap_bytes), read_only=True, data_only=True)
                sheet_sel = st.selectbox("Hoja", wb_ap.sheetnames, key="cf_ap_sheet")
                header_cells = next(wb_ap[sheet_sel].iter_rows(min_row=1, max_row=1))
                headers = [str(c.value).strip() if c.value else "" for c in header_cells]
                headers = [h for h in headers if h]
                c1, c2, c3 = st.columns(3)
                with c1:
                    vendor_col = st.selectbox("Columna con el proveedor", headers, key="cf_ap_vendor_col")
                with c2:
                    ref_col = st.selectbox("Columna con la factura/referencia", headers, key="cf_ap_ref_col")
                with c3:
                    match_col = st.selectbox("Columna a traer (PO / Recepción)", headers, key="cf_ap_match_col")

                if st.button("🔄 Cruzar ahora", type="primary", key="cf_ap_cross"):
                    matched, warnings = _cf_cross_reference_ap(ap_bytes, sheet_sel, vendor_col, ref_col, match_col)
                    for w in warnings:
                        st.error(w)
                    if matched:
                        st.success(f"✅ {matched} facturas actualizadas con su PO/Recepción")
                        st.rerun()
                    elif not warnings:
                        st.info("No se encontraron coincidencias nuevas.")
            except Exception as e:
                st.error(f"No se pudo leer el archivo: {e}")
            finally:
                del ap_bytes

    # ── Importar tu Excel actual ─────────────────────────────────────────────
    with cf_tabs[7]:
        st.subheader("Importar desde tu Excel actual (CONTROL_FACTURES)")
        st.caption(
            "Migra el contenido de tu archivo .xlsm/.xlsx actual (hojas CONTROL, FOURNISSEUR "
            "y cualquier hoja de reglas especiales tipo 'Xerox') hacia esta base, incluyendo "
            "los responsables y estados ya usados. Podés volver a ejecutarlo para reemplazar "
            "todo con una versión más nueva del Excel."
        )
        legacy_up = st.file_uploader("Tu archivo Excel actual", type=["xlsm", "xlsx"], key="cf_legacy_uploader")
        if legacy_up and st.button("📥 Importar", type="primary", key="cf_legacy_import_btn"):
            try:
                result = _cf_import_legacy_excel(legacy_up.read())
                st.success(
                    f"✅ Importado: {result['facturas']} facturas, {result['proveedores']} proveedores, "
                    f"{result['reglas']} reglas especiales, {result['responsables']} responsables nuevos, "
                    f"{result['estados']} estados nuevos."
                )
                for w in result["warnings"]:
                    st.warning(w)
                st.info(
                    "ℹ️ La columna **PAYÉE** del Excel original tenía fórmulas rotas (#REF!) en todas "
                    "las filas, así que se importó vacía — marcá manualmente las que ya estén pagadas. "
                    "Revisá también las columnas CC/GL importadas: algunas filas del archivo original "
                    "estaban incompletas o con datos inconsistentes."
                )
                st.rerun()
            except Exception as e:
                st.error(f"Error al importar: {e}")

